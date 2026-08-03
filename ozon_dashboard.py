"""
OZON P&L Dashboard
==================
Запуск: streamlit run ozon_dashboard.py
"""

import streamlit as st
import requests
import pandas as pd
from datetime import datetime, date, timedelta
import calendar

# streamlit-aggrid — для страницы «P&L» с раскрывающимися группами столбцов (Task #20).
# Импорт "мягкий": если пакет ещё не подхватился на Streamlit Cloud (requirements.txt
# обновлён, но деплой не прогнан) — страница «P&L» покажет обычную таблицу без групп
# столбцов вместо падения всего приложения.
try:
    from st_aggrid import AgGrid, JsCode
    _AGGRID_AVAILABLE = True
except Exception:
    _AGGRID_AVAILABLE = False

# Версия сборки — время последнего деплоя (проставляется вручную перед каждым git push,
# см. блок деплоя в OZON_DASHBOARD_CONTEXT.md). Показывается в сайдбаре, чтобы можно было
# на глаз проверить, подхватился ли последний пуш, не заходя на GitHub.
APP_BUILD_VERSION = "03.08.2026 23:15"

# ── Настройки страницы ──────────────────────────────────────────────────────
st.set_page_config(
    page_title="Ozon P&L",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
    .metric-box {
        background: #1a1d23;
        border: 1px solid #2a2d35;
        border-radius: 10px;
        padding: 16px 20px;
    }
    .stDataFrame { font-size: 13px; }
    div[data-testid="metric-container"] {
        background: #1a1d23;
        border: 1px solid #2a2d35;
        border-radius: 10px;
        padding: 12px 16px;
    }
</style>
""", unsafe_allow_html=True)

# ── Константы ───────────────────────────────────────────────────────────────
TAX       = 0.07
ACQUIRING = 0.015  # резервная ставка для отчёта реализации; для транзакций берётся из API (type_id=29)
AGENT_FEE = 20   # агентское для экспресс/FBS вместо последней мили

def ru_float(v, dec: int = 2) -> str:
    """1 234,56"""
    try:
        fv = float(v)
        if fv != fv:
            return "—"
        s = f"{abs(fv):,.{dec}f}"
        int_part, *dec_part = s.split(".")
        int_part = int_part.replace(",", " ")
        result = int_part + ("," + dec_part[0] if dec_part and dec > 0 else "")
        return ("-" if fv < 0 else "") + result
    except Exception:
        return str(v)

def ru_rub(v, dec: int = 0) -> str:
    """404,3 rub"""
    try:
        fv = float(v)
        if fv != fv:
            return "—"
        return ru_float(v, dec) + " ₽"
    except Exception:
        return str(v)

def ru_pct(v, dec: int = 1) -> str:
    """14,9%"""
    try:
        fv = float(v)
        if fv != fv:
            return "—"
        return f"{fv:.{dec}f}".replace(".", ",") + "%"
    except Exception:
        return str(v)

def r(n: float) -> str:
    """Рубли с пробелом как разделителем тысяч: 33 520 ₽"""
    return f"{n:,.0f} ₽".replace(",", " ")


# fee_type_id → читаемое название (item_fees и delivery.services)
# Перевод английских названий типов начислений (Ozon API возвращает en-названия)
_ACCRUAL_EN_TO_RU: dict[str, str] = {
    "PremiumSubscription":          "Подписка Premium",
    "Promotion":                    "Продвижение / Реклама",
    "PayPerClick":                  "Оплата за клик",
    "AcceleratedReviewCollection":  "Ускоренный сбор отзывов",
    "ItemCompensation":             "Компенсация за товар",
    "Placements":                   "Размещение на складе (платное хранение)",
    "DefectFineErrors":             "Штраф (ошибки)",
    "DefectFineShipmentDelayRate":  "Штраф (задержка отгрузки)",
    "StorageFee":                   "Хранение на складе",
    "Acquiring":                    "Эквайринг",
    "Installment":                  "Рассрочка",
    "Logistics":                    "Логистика",
    "Return":                       "Возврат",
    "Sale":                         "Продажа",
}

TYPE_NAMES: dict[int, str] = {
    1:  "Эквайринг",
    3:  "Реклама / Продвижение бренда",
    12: "Кросс-докинг",
    16: "Обработка Drop-off (ПВЗ)",
    17: "Обработка Drop-off партнёрами (ПВЗ)",
    22: "Рассрочка Ozon",
    29: "Доставка до ПВЗ партнёрами",
    32: "Логистика FBO",
    41: "Оплата за клик",
    45: "Обработка возвратов партнёрами",
    46: "Размещение на складе",
    48: "Бонусы продавца",
    52: "Подписка Premium / Premium Plus",
    54: "Продвижение с оплатой за заказ",
    59: "Обратная логистика",
    74: "Звёздные товары",
    77: "Поштучная приёмка",
    79: "Временное размещение партнёрами",
    93: "Штраф (индекс ошибок)",
    96: "Ускоренный сбор отзывов",
    98: "Доставка до ПВЗ силами Ozon",
}

# Категории для item_fees
ACQUIRING_TYPE_ID   = 1
PROMO_TYPE_IDS      = {3, 74}       # реклама по SKU → колонка "Реклама"
INSTALLMENT_TYPE_ID = 22

# ── Performance API (реклама CPC/CPO по артикулам) ───────────────────────────
# Seller API отдаёт расход на "Оплата за клик" (type_id 41) и "Продвижение с оплатой
# за заказ" (type_id 54) ОДНОЙ суммой на весь магазин (non_item_fee — см. STORE_COST_GROUPS
# ниже), без разбивки по SKU. Performance API (эндпоинт /api/client/statistics/products/sku)
# даёт расход по конкретным SKU за период. По решению пользователя (02.08.2026) —
# когда Performance API подключён, эта сумма ЗАМЕНЯЕТ 41+54 в «Расходах магазина»
# (а не складывается с ними), и распределяется по артикулам в основной таблице.
# Медийная/баннерная реклама (type_id 3, "Реклама / Продвижение бренда") сюда НЕ входит —
# она уже приходит per-SKU напрямую из Seller API (item_fees) и не нуждается в замене.
PERFORMANCE_API_URL = "https://api-performance.ozon.ru"
# В Ozon это ДВА структурно разных типа рекламы, с разными эндпоинтами статистики:
#  - CPC ("Трафареты", advObjectType=SKU) → POST /api/client/statistics/products/sku
#  - CPO ("Оплата за заказ", advObjectType=SEARCH_PROMO/ALL_SKU_PROMO) → отдельный
#    асинхронный отчёт /api/client/statistics/all_sku_promo/orders/generate
#    (products/sku эту модель не покрывает — подтверждено на реальном кабинете 02.08.2026:
#    кампания "Оплата за заказ - все товары" не давала расход через products/sku).
# type_id 41 (Seller API, "Оплата за клик") = CPC, type_id 54 ("Продвижение с оплатой
# за заказ") = CPO — исключаются из "Расходов магазина" НЕЗАВИСИМО друг от друга,
# только если соответствующий источник реально отдал данные (см. cpc_ok/cpo_ok ниже).
PERFORMANCE_CPC_STORE_TYPE_ID = 41
PERFORMANCE_CPO_STORE_TYPE_ID = 54
PERFORMANCE_STORE_TYPE_IDS = {PERFORMANCE_CPC_STORE_TYPE_ID, PERFORMANCE_CPO_STORE_TYPE_ID}

# non_item_fee — расходы магазина (не на артикул, а на кабинет в целом)
# Группы для отображения в блоке "Расходы магазина"
STORE_COST_GROUPS: dict[int, str] = {
    52: "Подписки",
    54: "Реклама",
    41: "Реклама",
    96: "Реклама",
    48: "Реклама",
    12: "Услуги FBO",
    46: "Услуги FBO",
    77: "Услуги FBO",
    79: "Услуги партнёров",
    93: "Штрафы",
}

EXPRESS_TARIFF = [(2000, 300), (4000, 400), (7500, 500), (20000, 600), (float("inf"), 800)]

def express_cost(order_total: float) -> float:
    for threshold, cost in EXPRESS_TARIFF:
        if order_total <= threshold:
            return cost
    return 800

def _get_type_id(obj: dict):
    """Читает type_id / accrual_id — Ozon переименовал поле 09.06.2026."""
    v = obj.get("accrual_id")
    if v is None:
        v = obj.get("type_id")
    return v

def collect_store_costs(ops: list[dict]) -> dict[int, float]:
    """Собирает non_item_fee по type_id/accrual_id — расходы уровня магазина."""
    totals: dict[int, float] = {}
    for accrual in ops:
        nif = accrual.get("non_item_fee")
        if not isinstance(nif, dict):
            continue
        tid = _get_type_id(nif)
        if tid is None:
            continue
        amt = float(((nif.get("accrued") or {}).get("amount") or 0))
        totals[tid] = totals.get(tid, 0) + amt
    return totals

def collect_store_costs_daily(ops: list[dict]) -> dict[str, dict[int, float]]:
    """
    То же, что collect_store_costs, но с разбивкой по дням (accrual["_date"]) —
    нужно для страницы «P&L» (Task #20), чтобы распределить расходы магазина
    (подписки/реклама-магазин/FBO/партнёры/штрафы) по неделям/месяцам, а не
    показывать одной суммой за весь период, как раньше в блоке «Расходы магазина».
    """
    totals: dict[str, dict[int, float]] = {}
    for accrual in ops:
        if not isinstance(accrual, dict):
            continue
        nif = accrual.get("non_item_fee")
        if not isinstance(nif, dict):
            continue
        tid = _get_type_id(nif)
        if tid is None:
            continue
        d = accrual.get("_date", "")
        amt = float(((nif.get("accrued") or {}).get("amount") or 0))
        totals.setdefault(d, {})
        totals[d][tid] = totals[d].get(tid, 0) + amt
    return totals

# ── Ozon API ────────────────────────────────────────────────────────────────
API_URL = "https://api-seller.ozon.ru"

def _secret(key: str, default: str = "") -> str:
    """
    Читает значение из Secrets приложения (Manage app → Settings → Secrets на
    Streamlit Cloud), если оно там задано — чтобы не вводить API-ключи заново
    при каждом открытии. Если Secrets не настроены — просто пусто, поля в
    сайдбаре остаются обычными пустыми text_input, как раньше.
    """
    try:
        return str(st.secrets.get(key, default))
    except Exception:
        return default

def api_post(endpoint: str, body: dict, client_id: str, api_key: str, _retries: int = 6) -> dict:
    """
    Запрос к Ozon Seller API.
    При 429 (rate limit) — тихо ждём и повторяем (до _retries раз, с нарастающей паузой),
    вместо того чтобы сразу показывать ошибку пользователю. Ozon режет по запросам в секунду,
    и при плотных сериях вызовов (диагностика, маппинг SKU, постраничная загрузка) это
    ожидаемая, а не аварийная ситуация — добавлено 02.08.2026.
    """
    headers = {
        "Client-Id": client_id,
        "Api-Key":   api_key,
        "Content-Type": "application/json",
    }
    import time as _time
    delay = 0.5
    for attempt in range(_retries + 1):
        try:
            r = requests.post(API_URL + endpoint, json=body, headers=headers, timeout=30)
        except Exception as e:
            if attempt < _retries:
                _time.sleep(delay)
                delay *= 2
                continue
            st.error(f"Ошибка соединения с API: {e}")
            return {}

        if r.status_code == 429:
            if attempt < _retries:
                _time.sleep(delay)
                delay *= 2
                continue
            st.error(f"Ошибка API 429 на {endpoint}: лимит запросов исчерпан, повторные попытки не помогли ({r.text[:200]})")
            return {}

        if r.status_code != 200:
            st.error(f"Ошибка API {r.status_code} на {endpoint}: {r.text[:300]}")
            return {}

        return r.json()

    return {}

@st.cache_data(ttl=300, show_spinner=False)
def fetch_realization(client_id: str, api_key: str, year: int, month: int) -> list[dict]:
    """
    /v2/finance/realization — месячный отчёт о реализации.
    Доступен только за предыдущие месяцы (Ozon формирует в начале след. месяца).
    Возвращает список строк по артикулам.
    """
    data = api_post(
        "/v2/finance/realization",
        {"year": year, "month": month},
        client_id, api_key,
    )
    result = data.get("result") if data else {}
    return (result or {}).get("rows", [])

@st.cache_data(ttl=300, show_spinner=False)
def fetch_transactions(client_id: str, api_key: str, date_from: str, date_to: str) -> list[dict]:
    """
    /v1/finance/accrual/by-day — новый API начислений.
    Запрашиваем по одному дню за раз, обходим весь период date_from..date_to.
    """
    all_accruals = []

    from_dt = date.fromisoformat(date_from)
    to_dt   = date.fromisoformat(date_to)
    cur     = from_dt

    progress_bar = st.progress(0)
    total_days = (to_dt - from_dt).days + 1
    day_count = 0

    import time as _time

    while cur <= to_dt:
        day_str = cur.strftime("%Y-%m-%d")
        last_id = ""

        while True:
            try:
                data = api_post(
                    "/v1/finance/accrual/by-day",
                    {"date": day_str, "last_id": last_id},
                    client_id, api_key,
                )

                if not data:
                    break

                accruals = data.get("accruals") or []
                if accruals:
                    for a in accruals:
                        if isinstance(a, dict):
                            a["_date"] = day_str
                    all_accruals.extend(accruals)

                last_id = data.get("last_id") or ""
                if not last_id or not accruals:
                    break

                # ВАЖНО (найдено 02.08.2026): если за день несколько страниц (last_id),
                # раньше эти запросы шли подряд БЕЗ паузы — пауза 0.15с стояла только между
                # днями, а не между страницами внутри дня. При большом объёме заказов в день
                # это и было источником 429. Пауза нужна перед КАЖДЫМ запросом, включая
                # продолжение пагинации.
                _time.sleep(0.15)

            except Exception as e:
                st.warning(f"Ошибка при загрузке дня {day_str}: {e}")
                break

        cur += timedelta(days=1)
        day_count += 1
        progress_bar.progress(min(day_count / total_days, 1.0))
        _time.sleep(0.15)  # пауза между днями — избегаем 429

    progress_bar.empty()
    return all_accruals

@st.cache_data(ttl=86400, show_spinner=False)
def _fetch_accrual_types_raw(client_id: str, api_key: str) -> dict[int, str]:
    """
    /v1/finance/accrual/types — справочник всех type_id с АНГЛИЙСКИМИ названиями (сырые,
    без перевода). Кэшируется на сутки — сам список типов у Ozon меняется редко.
    Перевод НЕ делаем здесь специально: если исправить опечатку/неверный перевод в
    _ACCRUAL_EN_TO_RU, это должно применяться сразу, без риска словить старый закэшированный
    результат перевода (найдено 02.08.2026 — «Placements» показывал старый неверный перевод
    даже после фикса словаря, пока перевод был закэширован вместе с самим fetch).
    """
    data = api_post("/v1/finance/accrual/types", {}, client_id, api_key)
    result = {}
    # API возвращает "accrual_types" (не "types")
    items = data.get("accrual_types") or data.get("types") or []
    for t in items:
        tid = t.get("accrual_id") or t.get("type_id") or t.get("id")
        name = t.get("name") or t.get("title") or t.get("description")
        if tid is not None and name:
            result[int(tid)] = str(name)
    return result

def fetch_accrual_types(client_id: str, api_key: str) -> dict[int, str]:
    """
    /v1/finance/accrual/types — справочник type_id → русское название.
    Сырой список (англ.) кэшируется на сутки, а перевод через _ACCRUAL_EN_TO_RU
    применяется КАЖДЫЙ РАЗ заново (не кэшируется) — см. docstring _fetch_accrual_types_raw.
    """
    raw = _fetch_accrual_types_raw(client_id, api_key)
    return {tid: _ACCRUAL_EN_TO_RU.get(name, name) for tid, name in raw.items()}

# Модульный (процесс-wide) таймер последнего запроса к /v1/analytics/turnover/stocks.
# ВАЖНО: лимит 1 запрос/мин — на весь Client-Id, а не на один вызов fetch_offer_ids_by_sku().
# Если пользователь грузит один период, а следом сразу другой (например, август → июль),
# это ДВА РАЗНЫХ вызова функции (разные наборы SKU → разные ключи кэша st.cache_data),
# и без общего таймера второй вызов ловит 429, даже если каждый по отдельности шлёт всего
# один запрос. Список используется вместо простой переменной, чтобы значение переживало
# переопределения модуля при hot-reload в Streamlit.
_turnover_stocks_last_call_ts = [0.0]

def _wait_turnover_stocks_cooldown():
    import time as _t
    elapsed = _t.time() - _turnover_stocks_last_call_ts[0]
    remaining = 61 - elapsed
    if remaining > 0:
        st.info(
            f"Ждём {remaining:.0f} сек — у /v1/analytics/turnover/stocks лимит 1 запрос/мин "
            f"на аккаунт (недавно уже был запрос, возможно из другой загрузки/периода)…"
        )
        _t.sleep(remaining)
    _turnover_stocks_last_call_ts[0] = _t.time()

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_offer_ids_by_sku(client_id: str, api_key: str, skus: tuple) -> dict:
    """
    Маппинг числовых SKU → offer_id продавца.
    Сначала пробует /v1/analytics/turnover/stocks (Product read-only),
    затем /v1/product/info для каждого SKU как запасной вариант.
    """
    if not skus:
        return {}
    result = {}

    # Попытка 1: /v1/analytics/turnover/stocks — принимает список SKU, возвращает offer_id.
    # КРИТИЧНО (найдено 02.08.2026 в официальной документации): у этого метода лимит
    # «не больше 1 запроса в минуту по одному Client-Id» — совсем не как у большинства
    # других методов (обычно 50 запросов/сек). Раньше код бил список SKU на батчи по 500
    # и слал их ПОДРЯД без паузы — второй батч почти всегда получал 429, потому что
    # 60 секунд ещё не прошло. Метод принимает до 1000 SKU за ОДИН запрос, так что почти
    # всегда достаточно одного вызова. Если SKU больше 1000 — ждём реальную минуту между
    # батчами (а не короткий backoff из api_post, который тут бесполезен).
    #
    # ВАЖНО: этот отчёт покрывает только SKU с оборачиваемостью/остатками за период — товары
    # с нулевым остатком/редкими продажами он может не вернуть вообще. Раньше код при ЛЮБОМ
    # непустом результате сразу возвращал его (`if result: return result`) и НИКОГДА не пытался
    # доматчить оставшиеся SKU через /v1/product/info — из-за этого часть артикулов
    # отображалась как «сырой» числовой SKU вместо реального артикула. Исправлено 02.08.2026:
    # теперь для всех SKU, которых не нашлось в turnover/stocks, всегда идём в fallback.
    def _call_turnover_stocks(batch: list[str]) -> dict:
        """
        Вызов turnover/stocks с РЕАКТИВНЫМ ожиданием на случай, если проактивный
        cooldown (_wait_turnover_stocks_cooldown) всё равно не спас от 429 — например,
        если лимит Ozon считается не «60 сек с момента последнего успешного запроса»,
        а как-то иначе (скользящее окно, штраф за серию предыдущих 429 и т.п. — это
        не подтверждено документацией, поэтому подстраховываемся). В отличие от общего
        api_post() (короткий backoff до ~7.5 сек, бесполезный для лимита 1/мин), здесь
        при 429 ждём честную минуту и пробуем ещё раз — до 2 попыток всего.
        """
        import time as _t
        headers = {"Client-Id": client_id, "Api-Key": api_key, "Content-Type": "application/json"}
        body = {"sku": batch, "limit": len(batch), "offset": 0}
        for attempt in range(2):
            _wait_turnover_stocks_cooldown()
            try:
                resp = requests.post(API_URL + "/v1/analytics/turnover/stocks",
                                      json=body, headers=headers, timeout=30)
            except Exception:
                return {}
            if resp.status_code == 429:
                if attempt == 0:
                    st.info("Ozon всё ещё отвечает 429 на turnover/stocks — ждём полную минуту и пробуем ещё раз…")
                    _t.sleep(65)
                    continue
                st.warning(
                    "turnover/stocks недоступен из-за лимита запросов после повторной попытки — "
                    "часть артикулов будет доматчена через запасные методы (могут занять дольше)."
                )
                return {}
            if resp.status_code != 200:
                return {}
            return resp.json()
        return {}

    sku_strings = [str(s) for s in skus]
    try:
        for i in range(0, len(sku_strings), 1000):
            batch = sku_strings[i:i+1000]
            data = _call_turnover_stocks(batch)
            if not data:
                continue
            for item in (data.get("items") or []):
                if not isinstance(item, dict):
                    continue
                offer_id = str(item.get("offer_id") or "")
                sku_val  = str(item.get("sku") or "")
                if offer_id and sku_val:
                    result[sku_val] = offer_id
    except Exception:
        pass

    def _api_post_silent(endpoint: str, body: dict, _retries: int = 3) -> dict:
        """Как api_post(), но без st.error — для best-effort догоматчинга, где часть
        запросов ожидаемо не найдёт товар (устаревший SKU, архив и т.п.). При 429 тоже
        тихо ждёт и повторяет — иначе под рейт-лимитом маппинг массово проваливался бы
        молча, даже когда повтор мог бы сработать."""
        import time as _time
        headers = {"Client-Id": client_id, "Api-Key": api_key, "Content-Type": "application/json"}
        delay = 0.5
        for attempt in range(_retries + 1):
            try:
                resp = requests.post(API_URL + endpoint, json=body, headers=headers, timeout=30)
            except Exception:
                return {}
            if resp.status_code == 429 and attempt < _retries:
                _time.sleep(delay)
                delay *= 2
                continue
            if resp.status_code != 200:
                return {}
            return resp.json()
        return {}

    missing = [s for s in sku_strings if s not in result]

    # Попытка 2: /v3/product/info/list — пакетный метод по идентификаторам (пришёл на смену
    # /v2/product/list, который Ozon отключил в феврале 2025). Схема запроса не была
    # подтверждена по актуальной документации (не удалось её прочитать), поэтому обёрнуто
    # в try/except «на всякий случай» — если формат не подойдёт, просто перейдём к попытке 3.
    if missing:
        try:
            for i in range(0, len(missing), 100):
                batch = [int(s) for s in missing[i:i+100]]
                data = _api_post_silent("/v3/product/info/list", {"sku": batch})
                items = ((data or {}).get("items") or []) if isinstance(data, dict) else []
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    offer_id = str(item.get("offer_id") or "")
                    sku_val  = str(item.get("sku") or "") or str(item.get("id") or "")
                    if offer_id and sku_val:
                        result[sku_val] = offer_id
        except Exception:
            pass

    # Попытка 3: /v1/product/info — по одному SKU, ТОЛЬКО для тех, что всё ещё не нашлись.
    # Молча (без st.error): этот метод, похоже, устарел и на многие SKU отвечает 404 —
    # это ожидаемо для части каталога, не повод показывать пользователю стену ошибок.
    still_missing = [s for s in sku_strings if s not in result]
    for sku_str in still_missing[:300]:  # лимит 300, чтобы не перегружать API при больших каталогах
        data = _api_post_silent("/v1/product/info", {"sku": int(sku_str)})
        if data:
            item = data.get("result") or data
            offer_id = str(item.get("offer_id") or "")
            if offer_id:
                result[sku_str] = offer_id

    return result

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_sku_map(client_id: str, api_key: str) -> dict:
    """
    Строит словарь {sku: offer_id} из заказов FBO и FBS.
    """
    sku_map = {}
    today = date.today()
    since = (today - timedelta(days=90)).strftime("%Y-%m-%d")
    to    = today.strftime("%Y-%m-%d")

    for schema, path in [("fbo", "/v3/posting/fbo/list"), ("fbs", "/v4/posting/fbs/list")]:
        try:
            offset = 0
            page_size = 100  # v3/fbo и v4/fbs оба имеют max=100
            while True:
                body = {
                    "dir": "ASC",
                    "filter": {
                        "since": since + "T00:00:00.000Z",
                        "to":    to    + "T23:59:59.000Z",
                    },
                    "limit": page_size, "offset": offset,
                    "with": {"financial_data": False, "analytics_data": False}
                }
                resp = api_post(path, body, client_id, api_key)

                if not resp:
                    break

                if schema == "fbo":
                    result = resp.get("result", {})
                    postings = result if isinstance(result, list) else result.get("postings", [])
                else:
                    result = resp.get("result", {})
                    postings = result.get("postings", []) if result else []

                if not postings:
                    break

                for p in postings:
                    if not isinstance(p, dict):
                        continue
                    for prod in (p.get("products") or []):
                        if not isinstance(prod, dict):
                            continue
                        sku = str(prod.get("sku") or "")
                        offer_id = str(prod.get("offer_id") or "")
                        if sku and offer_id and sku not in sku_map:
                            sku_map[sku] = offer_id

                if len(postings) < page_size:
                    break
                offset += len(postings)
        except Exception:
            pass

    return sku_map

TURNOVER_GRADE_RU = {
    "GRADES_NONE":     "⏳ Ожидается поставка",
    "GRADES_NOSALES":  "⚪ Нет продаж",
    "GRADES_GREEN":    "🟢 Хорошо",
    "GRADES_YELLOW":   "🟡 Средне",
    "GRADES_RED":      "🔴 Плохо",
    "GRADES_CRITICAL": "🔴 Критично",
}

STOCK_GRADE_EMOJI = {
    "GRADES_GREEN":    "🟢",
    "GRADES_YELLOW":   "🟡",
    "GRADES_RED":      "🔴",
    "GRADES_CRITICAL": "🔴",
    "GRADES_NOSALES":  "⚪",
    "GRADES_NONE":     "⏳",
}

@st.cache_data(ttl=300, show_spinner=False)
def fetch_stocks(client_id: str, api_key: str, skus: tuple) -> list[dict]:
    """
    /v1/analytics/turnover/stocks — оборачиваемость и остатки по SKU (от Ozon).
    skus — кортеж числовых SKU (из sku_map_cache).
    Возвращает список с полями: offer_id, name, sku, current_stock, ads, idc, idc_grade, turnover, turnover_grade.
    """
    if not skus:
        return []
    all_items = []
    sku_list = list(skus)
    batch_size = 1000
    offset = 0
    while True:
        batch = sku_list[offset: offset + batch_size]
        if not batch:
            break
        body = {"sku": [str(s) for s in batch], "limit": len(batch), "offset": 0}
        data = api_post("/v1/analytics/turnover/stocks", body, client_id, api_key)
        if not data:
            break
        items = data.get("items") or []
        all_items.extend(items)
        if len(items) < batch_size:
            break
        offset += batch_size
    return all_items

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_name_map(client_id: str, api_key: str) -> dict:
    """Строит {sku: название} из заказов FBO+FBS за 90 дней."""
    name_map = {}
    today = date.today()
    since = (today - timedelta(days=90)).strftime("%Y-%m-%d")
    to    = today.strftime("%Y-%m-%d")
    
    for schema, path in [("fbo", "/v3/posting/fbo/list"), ("fbs", "/v4/posting/fbs/list")]:
        try:
            offset = 0
            page_size = 100  # v3/fbo и v4/fbs оба имеют max=100
            while True:
                body = {
                    "dir": "ASC",
                    "filter": {
                        "since": since + "T00:00:00.000Z",
                        "to": to + "T23:59:59.000Z"
                    },
                    "limit": page_size,
                    "offset": offset,
                    "with": {"financial_data": False, "analytics_data": False}
                }
                resp = api_post(path, body, client_id, api_key)

                if not resp:
                    break

                result = resp.get("result", {})
                if schema == "fbo":
                    postings = result if isinstance(result, list) else result.get("postings", [])
                else:
                    postings = result.get("postings", []) if result else []

                if not postings:
                    break

                for p in postings:
                    if not isinstance(p, dict):
                        continue
                    for prod in (p.get("products") or []):
                        if not isinstance(prod, dict):
                            continue
                        sku = str(prod.get("sku") or "")
                        name = str(prod.get("name") or "")
                        if sku and name and sku not in name_map:
                            name_map[sku] = name

                if len(postings) < page_size:
                    break
                offset += len(postings)
        except Exception:
            pass
            
    return name_map

def _build_transaction_rows(ops: list[dict]) -> pd.DataFrame:
    """
    Общий разбор ответа /v1/finance/accrual/by-day в "плоские" строки — ОДНА строка
    на каждый product (POSTING) или каждый sku_fees.fee (ITEM), с колонкой "date"
    (день начисления, accrual["_date"]). Не группирует — группировку делают вызывающие
    функции (transactions_to_df — по sku/article, transactions_to_daily_pnl — по дате).
    Вынесено в отдельную функцию 03.08.2026, чтобы не дублировать парсинг JSON в двух
    местах и не рассинхронизировать логику при будущих фиксах полей.
    """
    if not ops:
        return pd.DataFrame()

    rows = []
    for accrual in ops:
        if not isinstance(accrual, dict):
            continue

        date_str = accrual.get("_date", "")

        # ── POSTING: продажи/возвраты с выручкой и комиссией ──────────────────
        posting = accrual.get("posting") or {}
        products = posting.get("products") or []

        for prod in products:
            if not isinstance(prod, dict):
                continue

            sku = str(prod.get("sku") or "")
            if not sku:
                continue
            # offer_id может приходить прямо из транзакционного API — используем его как артикул
            offer_id_direct = str(prod.get("offer_id") or "")

            comm_block = prod.get("commission") or {}
            delivery_block = prod.get("delivery") or {}

            def _amt(block: dict, key: str) -> float:
                sub = (block.get(key) or {}) if isinstance(block, dict) else {}
                return float((sub.get("amount") or 0) if isinstance(sub, dict) else 0)

            # sale_amount / seller_price — ПОЛНАЯ сумма заказа (Выручка + Программы партнёров + Баллы за скидки).
            # Используем ТОЛЬКО для расчёта qty (кол-во единиц в мультизаказе), это НЕ выручка.
            gross_val = _amt(comm_block, "sale_amount")
            seller_price_val = _amt(comm_block, "seller_price")

            # Настоящая Выручка (совпадает со столбцом "Выручка" в выгрузке Ozon "Отчёт по начислениям")
            revenue_val = _amt(comm_block, "sale_price")
            # Программы партнёров
            partner_val = _amt(comm_block, "coinvestment")
            # Баллы за скидки — реальные деньги от Ozon, но не облагаются налогом (по решению пользователя)
            bonus_val = _amt(comm_block, "bonus")

            # Комиссия Ozon
            commission_val = _amt(comm_block, "sale_commission")

            # Логистика = весь total_accrued (включает FBO + доставку партнёрами type_29 и др.)
            logi_block = (delivery_block.get("total_accrued") or {}) if isinstance(delivery_block, dict) else {}
            logistics_val = float((logi_block.get("amount") or 0) if isinstance(logi_block, dict) else 0)

            # Эквайринг — приходит через ITEM-начисления (item_fees), не из delivery.services
            acquiring_val = 0.0

            # Количество: gross (sale_amount) / seller_price (для мультизаказов)
            if seller_price_val > 0 and abs(gross_val) > 0:
                qty = max(1, round(abs(gross_val) / seller_price_val))
            else:
                qty = 1

            is_return = gross_val < 0
            is_sale = not is_return and gross_val != 0

            rows.append({
                "date": date_str,
                "sku": sku,
                "article": offer_id_direct if offer_id_direct else sku,  # offer_id приоритетнее SKU
                "name": "",
                "qty": qty if is_sale else 0,
                "qty_ret": qty if is_return else 0,
                "sale": revenue_val if is_sale else 0,
                "return": revenue_val if is_return else 0,
                "partner_programs": partner_val,
                "bonus_points": bonus_val,
                "commission": commission_val,
                "logistics": logistics_val,
                "acquiring":   acquiring_val,
                "promo":       0.0,
                "installment": 0.0,
                "other_costs": 0.0,
            })

        # ── ITEM: доп. сборы по SKU — эквайринг, реклама, прочее ────────────
        item_fees_block = accrual.get("item_fees") or {}
        for sku_fees in (item_fees_block.get("fees") or []):
            if not isinstance(sku_fees, dict):
                continue
            sku = str(sku_fees.get("sku") or "")
            if not sku:
                continue
            offer_id_fees = str(sku_fees.get("offer_id") or "")
            for fee in (sku_fees.get("fees") or []):
                if not isinstance(fee, dict):
                    continue
                tid = _get_type_id(fee)
                amount = float((fee.get("accrued") or {}).get("amount") or 0)
                known = {ACQUIRING_TYPE_ID, INSTALLMENT_TYPE_ID} | PROMO_TYPE_IDS
                rows.append({
                    "date": date_str,
                    "sku": sku, "article": offer_id_fees if offer_id_fees else sku, "name": "",
                    "qty": 0, "qty_ret": 0, "sale": 0, "return": 0,
                    "partner_programs": 0.0, "bonus_points": 0.0,
                    "commission": 0, "logistics": 0,
                    "acquiring":    amount if tid == ACQUIRING_TYPE_ID   else 0.0,
                    "promo":        amount if tid in PROMO_TYPE_IDS       else 0.0,
                    "installment":  amount if tid == INSTALLMENT_TYPE_ID  else 0.0,
                    "other_costs":  amount if tid not in known            else 0.0,
                })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    for col in ["qty", "qty_ret", "sale", "return", "partner_programs", "bonus_points",
                "commission", "logistics", "acquiring", "promo", "installment", "other_costs"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df

def transactions_to_df(ops: list[dict]) -> pd.DataFrame:
    """
    Разбираем ответ /v1/finance/accrual/by-day, группируем по артикулу (sku, article) —
    для таблицы «По артикулам» за весь выбранный период одной строкой на товар.
    """
    df = _build_transaction_rows(ops)
    if df.empty:
        return df

    grouped = df.groupby(["sku", "article"]).agg(
        name=("name", "first"),
        qty=("qty", "sum"),
        qty_ret=("qty_ret", "sum"),
        sale=("sale", "sum"),
        return_sum=("return", "sum"),
        partner_programs=("partner_programs", "sum"),
        bonus_points=("bonus_points", "sum"),
        commission=("commission", "sum"),
        logistics=("logistics", "sum"),
        acquiring=("acquiring", "sum"),
        promo=("promo", "sum"),
        installment=("installment", "sum"),
        other_costs=("other_costs", "sum"),
    ).reset_index()

    grouped["revenue"] = grouped["sale"] + grouped["return_sum"]
    return grouped

def transactions_to_daily_pnl(
    ops: list[dict], cost_map: dict,
    cpc_by_date: dict = None, cpo_by_date: dict = None,
) -> pd.DataFrame:
    """
    P&L по ДНЯМ (а не по артикулам) — для страницы «P&L» и для сохранения истории.
    Группирует по (date, sku, article), применяет себестоимость (нетто по qty_ret,
    как в enrich_with_cost) и рекламу CPC/CPO по дню+SKU (Performance API отдаёт дату
    в каждой строке — см. fetch_performance_sku_expense/fetch_all_sku_promo_orders_report),
    затем сворачивает до одной строки на календарный день.

    Считается ПОСЛЕ применения sku_map (article уже должен быть настоящим офер_id,
    не сырым числовым SKU) — передавай ops туда же, откуда взят df для transactions_to_df,
    но article в _build_transaction_rows берётся из самого accrual (offer_id), а не из
    sku_map, так что для товаров без offer_id в API артикул останется числовым SKU —
    это ожидаемо и не мешает суммам, только отображению в будущей детализации по товару.
    """
    fine = _build_transaction_rows(ops)
    if fine.empty:
        return fine

    fine["cost_price"] = fine["article"].map(cost_map).fillna(0)
    fine["qty_net"] = fine["qty"] - fine["qty_ret"]
    fine["cost_total"] = fine["cost_price"] * fine["qty_net"]

    cpc_by_date = cpc_by_date or {}
    cpo_by_date = cpo_by_date or {}

    daily = fine.groupby("date").agg(
        qty=("qty", "sum"),
        qty_ret=("qty_ret", "sum"),
        sale=("sale", "sum"),
        return_sum=("return", "sum"),
        partner_programs=("partner_programs", "sum"),
        bonus_points=("bonus_points", "sum"),
        commission=("commission", "sum"),
        logistics=("logistics", "sum"),
        acquiring=("acquiring", "sum"),
        promo=("promo", "sum"),
        installment=("installment", "sum"),
        other_costs=("other_costs", "sum"),
        cost_total=("cost_total", "sum"),
    ).reset_index()

    # Реклама по дням: считаем ОДИН РАЗ на уникальную пару (дата, SKU) из by_date-словарей
    # Performance API — НЕ по "плоским" строкам fine (там один и тот же SKU за один день
    # встречается по нескольку раз: отдельная строка от POSTING + отдельная от каждого
    # ITEM-сбора типа эквайринг/промо, иначе реклама задвоится/учетверится на каждую
    # такую строку). Заодно учитываем SKU без продаж в этот день (показы/клики без
    # покупки) — их расход всё равно суммируется в тот же день, деньги не теряются.
    def _daily_ads_totals(by_date: dict) -> dict:
        totals: dict[str, float] = {}
        for d, sku_map_ in by_date.items():
            totals[d] = totals.get(d, 0.0) + sum(sku_map_.values())
        return totals

    ads_cpc_totals = _daily_ads_totals(cpc_by_date)
    ads_cpo_totals = _daily_ads_totals(cpo_by_date)

    all_dates = set(daily["date"]) | set(ads_cpc_totals.keys()) | set(ads_cpo_totals.keys())
    daily = daily.set_index("date").reindex(sorted(all_dates)).reset_index()
    daily["ads_cpc"] = daily["date"].map(ads_cpc_totals).fillna(0.0)
    daily["ads_cpo"] = daily["date"].map(ads_cpo_totals).fillna(0.0)
    daily = daily.fillna(0)
    daily["revenue"] = daily["sale"] + daily["return_sum"]
    daily["partner_programs"] = daily.get("partner_programs", 0.0)
    daily["bonus_points"] = daily.get("bonus_points", 0.0)
    daily["total_income"] = daily["revenue"] + daily["partner_programs"] + daily["bonus_points"]
    daily["tax"] = TAX * (daily["revenue"] + daily["partner_programs"])
    daily["acquiring"] = daily["acquiring"].abs()
    daily["promo"] = daily["promo"].abs()
    daily["installment"] = daily["installment"].abs()
    daily["ads_perf"] = daily["ads_cpc"] + daily["ads_cpo"]
    daily["profit"] = (
        daily["total_income"]
        + daily["commission"]
        + daily["logistics"]
        - daily["promo"]
        - daily["installment"]
        + daily["other_costs"]
        - daily["cost_total"]
        - daily["acquiring"]
        - daily["tax"]
        - daily["ads_perf"]
    )
    daily["margin_pct"] = (daily["profit"] / daily["total_income"].replace(0, float("nan"))) * 100
    return daily.sort_values("date").reset_index(drop=True)

def build_period_pnl(
    daily: pd.DataFrame, store_daily: dict, freq: str,
    perf_replaces_cpc: bool = False, perf_replaces_cpo: bool = False,
) -> pd.DataFrame:
    """
    Сводка P&L по периодам (день/неделя/месяц) для страницы «P&L» (Task #20).

    ВАЖНО (эпистемика — это МОЯ интерпретация макета из скриншота стороннего сервиса,
    а не официальная методология Ozon): «Операционная прибыль» здесь = доход минус
    себестоимость, комиссия и расходы внутри МП (логистика/эквайринг/рассрочка/прочее),
    ДО вычета маркетинга и налога. «Чистая прибыль» = операционная минус маркетинг и
    налог. ROI считается от себестоимости (Себестоимость), как это принято в юнит-
    экономике маркетплейсов, но это тоже конвенция, а не единственно верный вариант —
    некоторые считают ROI от выручки. Если формулы должны быть другими — легко поменять
    здесь, в одном месте.

    daily         — результат transactions_to_daily_pnl (по дате, уже с ads_cpc/ads_cpo/ads_perf).
    store_daily   — результат collect_store_costs_daily(raw_ops) (non_item_fee по дате и type_id).
    freq          — "D" (день), "W" (неделя, с понедельника), "M" (месяц).
    perf_replaces_cpc/cpo — если Performance API реально заменил type_id 41/54 (см. основной
                     блок KPI) — тогда эти type_id ИСКЛЮЧАЮТСЯ из «Маркетинг (магазин)», чтобы
                     не задвоить с daily["ads_perf"] (которая уже считает CPC/CPO по SKU).
    """
    if daily is None or daily.empty:
        return pd.DataFrame()

    daily = daily.copy()
    daily["date"] = pd.to_datetime(daily["date"])

    # ── Разворачиваем store_daily в DataFrame по дате ────────────────────────
    excluded_marketing_tids = set()
    if perf_replaces_cpc:
        excluded_marketing_tids.add(PERFORMANCE_CPC_STORE_TYPE_ID)
    if perf_replaces_cpo:
        excluded_marketing_tids.add(PERFORMANCE_CPO_STORE_TYPE_ID)

    store_rows = []
    for d, tid_map in (store_daily or {}).items():
        marketing_store = 0.0
        other_store = 0.0
        for tid, amt in tid_map.items():
            group = STORE_COST_GROUPS.get(tid, "Прочее")
            if group == "Реклама" and tid in excluded_marketing_tids:
                continue  # уже учтено в ads_perf (Performance API)
            if group == "Реклама":
                marketing_store += amt
            else:
                other_store += amt
        store_rows.append({"date": d, "marketing_store": marketing_store, "other_store": other_store})

    store_df = pd.DataFrame(store_rows, columns=["date", "marketing_store", "other_store"])
    if not store_df.empty:
        store_df["date"] = pd.to_datetime(store_df["date"])

    # ── Объединяем: даты либо из daily, либо только из store_daily (напр. подписка
    # списана в день без продаж) ─────────────────────────────────────────────────
    all_dates = pd.Index(daily["date"])
    if not store_df.empty:
        all_dates = all_dates.union(pd.Index(store_df["date"]))
    base = pd.DataFrame({"date": sorted(all_dates)})

    merged = base.merge(daily, on="date", how="left").merge(store_df, on="date", how="left")
    num_cols = [c for c in merged.columns if c != "date"]
    merged[num_cols] = merged[num_cols].fillna(0.0)

    # ── Метка периода ────────────────────────────────────────────────────────
    if freq == "D":
        merged["period_start"] = merged["date"]
    elif freq == "W":
        merged["period_start"] = merged["date"] - pd.to_timedelta(merged["date"].dt.weekday, unit="D")
    elif freq == "M":
        merged["period_start"] = merged["date"].values.astype("datetime64[M]")
    else:
        raise ValueError(f"Неизвестная granularity: {freq}")

    agg_cols = [
        "qty", "qty_ret", "revenue", "total_income", "commission", "logistics",
        "acquiring", "promo", "installment", "other_costs", "cost_total",
        "ads_cpc", "ads_cpo", "ads_perf", "tax", "marketing_store", "other_store",
    ]
    agg_cols = [c for c in agg_cols if c in merged.columns]
    period = merged.groupby("period_start")[agg_cols].sum().reset_index()

    # ── Подписи периодов ─────────────────────────────────────────────────────
    def _label(ts) -> str:
        if freq == "D":
            return ts.strftime("%d.%m.%Y")
        if freq == "W":
            end = ts + pd.Timedelta(days=6)
            return f"{ts.strftime('%d.%m')} – {end.strftime('%d.%m.%Y')}"
        months_ru = ["январь","февраль","март","апрель","май","июнь","июль",
                     "август","сентябрь","октябрь","ноябрь","декабрь"]
        return f"{months_ru[ts.month - 1].capitalize()} {ts.year}"

    period["period_label"] = period["period_start"].apply(_label)

    # ── Продажи ──────────────────────────────────────────────────────────────
    period["qty_total"] = period["qty"] + period["qty_ret"]

    # ── Комиссия ─────────────────────────────────────────────────────────────
    period["commission_abs"] = period["commission"].abs()
    period["commission_pct"] = (period["commission_abs"] / period["revenue"].replace(0, float("nan"))) * 100

    # ── Расходы внутри МП (логистика/эквайринг/рассрочка/прочее по SKU + прочее магазина) ──
    # non_item_fee (marketing_store/other_store) приходят из Ozon отрицательными (списание),
    # как и в collect_store_costs() — берём abs(), т.к. все остальные "*_abs" тут уже в
    # положительных числах-расходах.
    period["logistics_abs"]   = period["logistics"].abs()
    period["other_costs_abs"] = period["other_costs"].abs()
    period["other_store_abs"] = period["other_store"].abs()
    period["mp_expenses_total"] = (
        period["logistics_abs"] + period["acquiring"] + period["installment"]
        + period["other_costs_abs"] + period["other_store_abs"]
    )

    # ── Маркетинг (по SKU-реклама из item_fees + Performance API CPC/CPO + магазинная реклама,
    # не заменённая Performance API) ────────────────────────────────────────────
    period["promo_abs"] = period["promo"].abs()
    period["marketing_store_abs"] = period["marketing_store"].abs()
    period["marketing_total"] = period["promo_abs"] + period["ads_perf"] + period["marketing_store_abs"]

    # ── Прибыль ──────────────────────────────────────────────────────────────
    period["operating_profit"] = (
        period["total_income"] - period["cost_total"] - period["commission_abs"] - period["mp_expenses_total"]
    )
    period["roi_operating_pct"] = (period["operating_profit"] / period["cost_total"].replace(0, float("nan"))) * 100
    period["net_profit"] = period["operating_profit"] - period["marketing_total"] - period["tax"]
    period["roi_net_pct"] = (period["net_profit"] / period["cost_total"].replace(0, float("nan"))) * 100

    return period.sort_values("period_start").reset_index(drop=True)

def apply_sku_map(df: pd.DataFrame, sku_map: dict, name_map: dict = None) -> pd.DataFrame:
    """Заменяет числовой SKU на твой артикул (offer_id) и название через справочник."""
    if df.empty or not sku_map:
        return df
    df = df.copy()
    df["article"] = df["sku"].map(sku_map).fillna(df["article"])
    if name_map:
        df["name"] = df["sku"].map(name_map).fillna(df["name"])
    return df

def realization_to_df(rows: list[dict]) -> pd.DataFrame:
    """
    Парсим ответ /v2/finance/realization.
    Каждая строка — одна продажа. Группируем по offer_id.
    """
    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue

        item = r.get("item") or {}
        dc = r.get("delivery_commission") or {}
        rc = r.get("return_commission")  # может быть null

        article = str(item.get("offer_id") or item.get("sku") or "—")
        name = str(item.get("name") or "")
        sku = str(item.get("sku") or "")

        qty_sold = int(dc.get("quantity") or 0)
        qty_ret = int((rc or {}).get("quantity") or 0)

        seller_price = float(r.get("seller_price_per_instance") or 0)
        revenue = seller_price * qty_sold
        return_amount = -(seller_price * qty_ret)

        # Комиссия Ozon = standard_fee (продажа + возврат)
        commission = -float(dc.get("standard_fee") or 0)
        if rc:
            commission += -float(rc.get("standard_fee") or 0)

        # Логистика = delivery_fee (доставка покупателю + доставка возврата)
        logistics = -float(dc.get("delivery_fee") or 0)
        if rc:
            logistics += -float(rc.get("delivery_fee") or 0)

        out.append({
            "article": article,
            "name": name,
            "sku": sku,
            "qty": qty_sold,
            "qty_ret": qty_ret,
            "revenue": revenue,
            "return_sum": return_amount,
            "commission": commission,
            "logistics": logistics,
            "acquiring":   0,
            "promo":       0,
            "installment": 0,
            "other_costs": 0,
        })

    if not out:
        return pd.DataFrame()

    df = pd.DataFrame(out)
    for col in ["qty", "qty_ret", "revenue", "return_sum", "commission", "logistics", "acquiring", "promo", "other_costs"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    grouped = df.groupby(["article", "sku"]).agg(
        name=("name", "first"),
        qty=("qty", "sum"),
        qty_ret=("qty_ret", "sum"),
        revenue=("revenue", "sum"),
        return_sum=("return_sum", "sum"),
        commission=("commission", "sum"),
        logistics=("logistics", "sum"),
        acquiring=("acquiring", "sum"),
        promo=("promo", "sum"),
        installment=("installment", "sum"),
        other_costs=("other_costs", "sum"),
    ).reset_index()

    return grouped

def enrich_with_cost(df: pd.DataFrame, cost_map: dict, cpc_ad_map: dict = None, cpo_ad_map: dict = None) -> pd.DataFrame:
    """
    Добавляем себестоимость, налог, эквайринг и считаем прибыль.

    cpc_ad_map / cpo_ad_map: {sku: расход в рублях} из Performance API — CPC (оплата
    за клик) и CPO (оплата за заказ) ОТДЕЛЬНО, чтобы в таблице по артикулам было видно
    каждую статью раздельно, а не одной смешанной суммой. Складываются в "ads_perf"
    (используется в расчёте прибыли, как раньше). Sku, которых нет в df (например,
    показы были, а продаж в периоде не было), сюда НЕ попадают — их сумма учитывается
    отдельно на уровне «Расходов магазина» (см. основной блок ниже), чтобы деньги не
    терялись из P&L.
    """
    if df.empty:
        return df

    df = df.copy()
    df["cost_price"] = df["article"].map(cost_map).fillna(0)
    # Себестоимость считаем НЕТТО: проданные минус возвращённые в том же периоде штуки
    # (решение пользователя, 02.08.2026) — возвращённый товар приходит обратно целым, его
    # себестоимость не списывается в убыток. qty_ret может отсутствовать (demo-данные) —
    # тогда просто 0, поведение не меняется.
    if "qty_ret" not in df.columns:
        df["qty_ret"] = 0
    df["qty_net"] = df["qty"] - df["qty_ret"]
    df["cost_total"] = df["cost_price"] * df["qty_net"]
    # Если эквайринг пришёл из API (транзакции) — используем его; иначе расчётный %
    if "acquiring" not in df.columns or df["acquiring"].abs().sum() == 0:
        df["acquiring"] = ACQUIRING * df["revenue"]
    else:
        df["acquiring"] = df["acquiring"].abs()  # приводим к положительному для отображения расхода

    # Программы партнёров и Баллы за скидки — из транзакционного API (посылка commission.coinvestment /
    # commission.bonus). Для отчёта о реализации (realization_to_df) этих колонок нет — считаем их нулевыми.
    if "partner_programs" not in df.columns:
        df["partner_programs"] = 0.0
    if "bonus_points" not in df.columns:
        df["bonus_points"] = 0.0

    # Налог УСН 7%: база = Выручка + Программы партнёров. Баллы за скидки в базу НЕ входят
    # (решение пользователя, 02.08.2026) — см. OZON_DASHBOARD_CONTEXT.md.
    df["tax"] = TAX * (df["revenue"] + df["partner_programs"])

    # Полный доход = Выручка + Программы партнёров + Баллы за скидки (реальные деньги,
    # которые платит Ozon и которые сходятся с "Итого к выплате" в кабинете Ozon).
    df["total_income"] = df["revenue"] + df["partner_programs"] + df["bonus_points"]

    # promo и installment приводим к положительному (расход), как acquiring
    for col in ("promo", "installment"):
        if col in df.columns:
            df[col] = df[col].abs()
        else:
            df[col] = 0.0
    if "other_costs" not in df.columns:
        df["other_costs"] = 0.0

    # Реклама CPC и CPO по артикулам (Performance API), отдельными колонками — только если
    # переданы карты. Если не переданы (Performance API не подключён / выключен галочкой) —
    # колонки нулевые, и эта статья расхода остаётся там же, где была раньше: одной суммой
    # в "Расходах магазина".
    df["ads_cpc"] = df["sku"].astype(str).map(cpc_ad_map).fillna(0.0) if cpc_ad_map else 0.0
    df["ads_cpo"] = df["sku"].astype(str).map(cpo_ad_map).fillna(0.0) if cpo_ad_map else 0.0
    df["ads_perf"] = df["ads_cpc"] + df["ads_cpo"]

    df["profit"] = (
        df["total_income"]
        + df["commission"]     # отрицательная
        + df["logistics"]      # отрицательная
        - df["promo"]          # положительная (расход) — после abs()
        - df["installment"]    # положительная (расход) — после abs()
        + df["other_costs"]    # прочие (могут быть ± )
        - df["cost_total"]
        - df["acquiring"]
        - df["tax"]
        - df["ads_perf"]       # реклама CPC/CPO по артикулу (Performance API)
    )
    df["margin_pct"] = (df["profit"] / df["total_income"].replace(0, float("nan"))) * 100

    # По запросу пользователя (02.08.2026): "Выручка" в интерфейсе должна быть ПОЛНОЙ, с баллами
    # за скидки и программами партнёров. tax и profit выше уже посчитаны на правильной базе
    # (revenue+partner_programs — без баллов), так что просто переопределяем колонку ПОСЛЕ расчётов.
    df["revenue"] = df["total_income"]
    return df

# ── Performance API (реклама CPC/CPO по артикулам, отдельная авторизация) ───
@st.cache_data(ttl=1500, show_spinner=False)
def get_performance_token(perf_client_id: str, perf_client_secret: str) -> str:
    """
    OAuth-токен Performance API (client_credentials). Токен живёт 1800 сек —
    кэшируем на 1500, чтобы Streamlit сам обновил его с запасом до истечения.
    Это ОТДЕЛЬНАЯ пара ключей от Client-ID/API-Key Seller API выше.
    """
    resp = requests.post(
        f"{PERFORMANCE_API_URL}/api/client/token",
        json={
            "client_id": perf_client_id,
            "client_secret": perf_client_secret,
            "grant_type": "client_credentials",
        },
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]

def fetch_performance_campaigns(token: str) -> list[dict]:
    """GET /api/client/campaign — все кампании кабинета, без фильтра."""
    resp = requests.get(
        f"{PERFORMANCE_API_URL}/api/client/campaign",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    resp.raise_for_status()
    return (resp.json() or {}).get("list", [])

def fetch_performance_sku_expense(
    token: str, campaign_ids: list[str], date_from: str, date_to: str,
) -> tuple[dict[str, float], dict[str, dict[str, float]], list[str], list[dict]]:
    """
    POST /api/client/statistics/products/sku — расход по SKU за период, ТОЛЬКО для
    CPC-кампаний (advObjectType=SKU, "Трафареты"/оплата за клик). Для CPO ("Оплата за
    заказ") этот метод пуст — см. fetch_all_sku_promo_orders_report ниже.
    По документации метод "не расходует лимиты Performance API".

    Запрос по списку campaignIds — ВСЁ ИЛИ НИЧЕГО: если хотя бы одна кампания в списке
    вызывает ошибку (например, архивная/некорректная), падает ВЕСЬ запрос, включая
    статистику по остальным нормальным кампаниям (проверено на реальном кабинете
    03.08.2026 — 28 кампаний упали одним запросом разом). Поэтому при ошибке батча
    рекурсивно делим его пополам ("бисекция"), пока не изолируем конкретные проблемные
    campaignId — так все "здоровые" кампании всё равно отдают расход, а не теряются
    вместе с одной сломанной. Метод не тратит лимиты Performance API — можно позволить
    себе лишние запросы на изоляцию без риска упереться в rate limit.

    Ответ содержит поле "date" по каждой строке — используем его, чтобы получить
    расход И по SKU суммарно за период (totals), И по дням (by_date, для P&L по
    периодам/истории), без второго запроса.

    Возвращает (sku → сумма расхода за весь период, {дата: {sku: расход}},
    список ID кампаний с ошибкой, список примеров реальных ошибок API — макс. 5 шт).
    """
    totals: dict[str, float] = {}
    by_date: dict[str, dict[str, float]] = {}
    failed: list[str] = []
    error_samples: list[dict] = []
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    def _query(ids: list[str]):
        """Один HTTP-запрос. Возвращает (ok, status, текст_ошибки)."""
        try:
            resp = requests.post(
                f"{PERFORMANCE_API_URL}/api/client/statistics/products/sku",
                json={"campaignIds": ids, "dateFrom": date_from, "dateTo": date_to},
                headers=headers, timeout=60,
            )
        except Exception as e:
            return False, None, str(e)
        if resp.status_code != 200:
            return False, resp.status_code, resp.text[:300]
        for row in (resp.json() or {}).get("rows", []):
            sku = str(row.get("sku") or "")
            if not sku:
                continue
            expense = float(row.get("expense") or 0)
            totals[sku] = totals.get(sku, 0.0) + expense
            d = str(row.get("date") or "")[:10]
            if d:
                by_date.setdefault(d, {})
                by_date[d][sku] = by_date[d].get(sku, 0.0) + expense
        return True, 200, None

    def _process(ids: list[str]):
        if not ids:
            return
        ok, status, text = _query(ids)
        if ok:
            return
        if len(ids) == 1:
            failed.extend(ids)
            if len(error_samples) < 5:
                error_samples.append({"campaignId": ids[0], "status": status, "error": text})
            return
        mid = len(ids) // 2
        _process(ids[:mid])
        _process(ids[mid:])

    BATCH = 50
    for i in range(0, len(campaign_ids), BATCH):
        _process(campaign_ids[i:i + BATCH])

    return totals, by_date, failed, error_samples

def fetch_cpc_report_async(
    token: str, campaign_ids: list[str], date_from: str, date_to: str, timeout_s: int = 90,
) -> dict:
    """
    Асинхронный отчёт по CPC-кампаниям ("Трафареты"/оплата за клик) за ЛЮБОЙ период —
    в отличие от fetch_performance_sku_expense (POST /api/client/statistics/products/sku),
    который, как выяснилось на реальном кабинете 03.08.2026, принимает только "today
    или yesterday" (в документации: dateFrom "не раньше предыдущего дня") — при попытке
    запросить более раннюю дату Ozon отвечает 400
    {"error":"date range must contain only today or yesterday"}. Это НЕ баг бисекции —
    это жёсткое ограничение самого метода, обойти батчами/повторами нельзя.

    Используем общий асинхронный эндпоинт POST /api/client/statistics/json (кампании
    любого типа, включая "Оплата за клик" — см. документацию, раздел "Архивы с
    примерами отчётов... Оплата за клик"), тот же поток generate → poll UUID → report,
    что уже работает для CPO (fetch_all_sku_promo_orders_report). Максимальный период
    одного запроса — 62 дня (задокументировано) → бьём диапазон на куски.

    ВАЖНО (честно про неопределённость — это НЕ проверено на реальных данных, в
    отличие от CPO-отчёта, который сверяли построчно с выгруженным файлом): точные
    ключи JSON-строк этого отчёта для CPC Ozon явно не документирует — только
    человекочитаемые названия колонок. Предполагаем ту же схему полей, что и у
    синхронного /statistics/products/sku (campaignId, sku, date, expense...), т.к.
    Ozon обычно переиспользует формат строк между sync/async версией одного отчёта —
    но это ПРЕДПОЛОЖЕНИЕ. Если формат другой — вернёт raw_sample с примером сырых
    строк вместо тихого нуля, чтобы можно было поправить парсинг за один проход.
    """
    out = {"by_sku": {}, "by_date": {}, "total": 0.0, "error": "", "raw_sample": None, "chunks_failed": []}
    if not campaign_ids:
        return out
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    from_dt = date.fromisoformat(date_from)
    to_dt = date.fromisoformat(date_to)
    chunks = []
    cur = from_dt
    while cur <= to_dt:
        chunk_end = min(cur + timedelta(days=61), to_dt)   # 62 дня включительно (задокументированный максимум)
        chunks.append((cur.isoformat(), chunk_end.isoformat()))
        cur = chunk_end + timedelta(days=1)

    import time as _time
    totals: dict[str, float] = {}
    by_date: dict[str, dict[str, float]] = {}
    raw_sample = None
    CAMPAIGN_BATCH = 100   # на случай недокументированного лимита на размер campaigns[]

    for chunk_from, chunk_to in chunks:
        for i in range(0, len(campaign_ids), CAMPAIGN_BATCH):
            batch_ids = campaign_ids[i:i + CAMPAIGN_BATCH]
            period_label = f"{chunk_from}..{chunk_to}"
            try:
                gen = requests.post(
                    f"{PERFORMANCE_API_URL}/api/client/statistics/json",
                    json={"campaigns": batch_ids, "dateFrom": chunk_from, "dateTo": chunk_to},
                    headers=headers, timeout=30,
                )
            except Exception as e:
                out["chunks_failed"].append({"period": period_label, "error": str(e)})
                continue
            if gen.status_code != 200:
                out["chunks_failed"].append({"period": period_label, "status": gen.status_code, "error": gen.text[:300]})
                continue
            uuid = (gen.json() or {}).get("UUID")
            if not uuid:
                out["chunks_failed"].append({"period": period_label, "error": f"нет UUID в ответе: {gen.text[:200]}"})
                continue

            state = None
            deadline = _time.time() + timeout_s
            while _time.time() < deadline:
                st_resp = requests.get(f"{PERFORMANCE_API_URL}/api/client/statistics/{uuid}", headers=headers, timeout=30)
                if st_resp.status_code != 200:
                    _time.sleep(3)
                    continue
                st_data = st_resp.json() or {}
                state = st_data.get("state")
                if state in ("OK", "ERROR"):
                    break
                _time.sleep(3)
            if state != "OK":
                out["chunks_failed"].append({"period": period_label, "error": f"отчёт не готов (статус: {state})"})
                continue

            rep = requests.get(
                f"{PERFORMANCE_API_URL}/api/client/statistics/report",
                params={"UUID": uuid}, headers=headers, timeout=60,
            )
            if rep.status_code != 200:
                out["chunks_failed"].append({"period": period_label, "error": f"report {rep.status_code}: {rep.text[:200]}"})
                continue

            rows = None
            try:
                parsed = rep.json()
                rows = parsed if isinstance(parsed, list) else (parsed.get("rows") or parsed.get("items"))
            except Exception:
                rows = None
            if rows is None:
                import csv, io
                text = rep.text
                delimiter = ";" if text.count(";") > text.count(",") else ","
                all_rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
                rows = [dict(zip([h.strip() for h in all_rows[0]], r)) for r in all_rows[1:]] if len(all_rows) >= 2 else []

            if rows and raw_sample is None:
                raw_sample = rows[:3]

            for row in rows or []:
                if not isinstance(row, dict):
                    continue
                sku = str(row.get("sku") or row.get("SKU") or "").strip()
                if not sku:
                    continue
                try:
                    expense = float(str(row.get("expense") or row.get("расход") or "0").replace(",", "."))
                except Exception:
                    expense = 0.0
                totals[sku] = totals.get(sku, 0.0) + expense
                d = str(row.get("date") or "").strip()[:10]
                if d:
                    by_date.setdefault(d, {})
                    by_date[d][sku] = by_date[d].get(sku, 0.0) + expense

    out["by_sku"] = totals
    out["by_date"] = by_date
    out["total"] = sum(totals.values())
    out["raw_sample"] = raw_sample
    if not totals and out["chunks_failed"]:
        out["error"] = f"Не удалось получить ни одного куска отчёта ({len(out['chunks_failed'])} ошибок за периоды) — см. диагностику."
    return out

def parse_all_sku_promo_orders_file(uploaded_file) -> tuple[dict[str, float], dict[str, dict[str, float]], str]:
    """
    Парсит файл 'Оплата за заказ (все товары). Отчёт по заказам', выгружаемый вручную
    из личного кабинета Ozon (xlsx/csv). Формат подтверждён на реальном файле
    пользователя 02.08.2026: первые 2 строки — период и заголовок отчёта, 3-я —
    заголовки колонок (Дата, ID заказа, Номер заказа, SKU, SKU продвигаемого товара,
    Артикул, Название товара, Количество, Стоимость продажи ₽, Стоимость ₽,
    Ставка %, Ставка ₽, Расход ₽).

    Берём колонку "SKU" (что реально продано), а НЕ "SKU продвигаемого товара"
    (может отличаться при кросс-показах) — расход ложится на тот же артикул,
    по которому в таблице уже считается выручка. Колонка "Дата" (формат ДД.ММ.ГГГГ
    в файле) используется для дневной разбивки — для P&L по периодам/истории.

    Возвращает (sku → сумма расхода за весь файл, {дата ГГГГ-ММ-ДД: {sku: расход}},
    текст ошибки или "" если всё ок).
    """
    try:
        name = (getattr(uploaded_file, "name", "") or "").lower()
        if name.endswith(".csv"):
            raw = pd.read_csv(uploaded_file, header=None, sep=None, engine="python")
        else:
            raw = pd.read_excel(uploaded_file, header=None)
    except Exception as e:
        return {}, {}, f"Не удалось прочитать файл: {e}"

    header_row_idx = None
    for i in range(min(10, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if "SKU" in vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        return {}, {}, "Не нашла строку с заголовками колонок (ищу ячейку 'SKU') — формат файла отличается от ожидаемого"

    headers = [str(v).strip() for v in raw.iloc[header_row_idx].tolist()]
    data = raw.iloc[header_row_idx + 1:].copy()
    data.columns = headers

    sku_col = next((c for c in headers if c == "SKU"), None)
    expense_col = next((c for c in headers if str(c).strip().lower().startswith("расход")), None)
    date_col = next((c for c in headers if str(c).strip().lower() == "дата"), None)
    if not sku_col or not expense_col:
        return {}, {}, f"Не нашла колонки 'SKU' / 'Расход, ₽'. Колонки в файле: {headers}"

    keep_cols = [sku_col, expense_col] + ([date_col] if date_col else [])
    data = data[keep_cols].dropna(subset=[sku_col])
    data[sku_col] = data[sku_col].astype(str).str.strip()
    data = data[data[sku_col].str.len() > 0]
    data[expense_col] = (
        data[expense_col].astype(str)
        .str.replace("\xa0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0.0)
    )

    grouped = data.groupby(sku_col)[expense_col].sum()
    totals = {str(k): float(v) for k, v in grouped.items()}

    by_date: dict[str, dict[str, float]] = {}
    if date_col:
        data["_iso_date"] = pd.to_datetime(data[date_col], dayfirst=True, errors="coerce").dt.strftime("%Y-%m-%d")
        for d, sub in data.dropna(subset=["_iso_date"]).groupby("_iso_date"):
            by_date[d] = sub.groupby(sku_col)[expense_col].sum().to_dict()

    return totals, by_date, ""

def fetch_all_sku_promo_orders_report(token: str, date_from: str, date_to: str, timeout_s: int = 90) -> dict:
    """
    Best-effort: асинхронный отчёт CPO 'Оплата за заказ — все товары. Отчёт по заказам'.
    Поток: GET .../all_sku_promo/orders/generate/json → UUID → поллинг статуса
    GET /api/client/statistics/{UUID} до state OK/ERROR → скачивание
    GET /api/client/statistics/report?UUID=...

    ВАЖНО (честно про неопределённость): точная JSON-схема СТРОК этого конкретного
    отчёта не задокументирована Ozon — в документации только человекочитаемые
    названия колонок, без английских ключей. Парсинг ниже ищет колонку SKU/расход
    по названию ключа гибко, а не по жёстко зашитым именам. Если формат окажется
    другим — вернёт понятную ошибку с примером сырых данных, а не тихий ноль.
    Надёжный запасной вариант — загрузка того же отчёта файлом (см. sidebar).
    """
    out = {"by_sku": {}, "by_date": {}, "total": 0.0, "error": "", "raw_sample": None}
    headers = {"Authorization": f"Bearer {token}"}
    tb_from, tb_to = f"{date_from}T00:00:00Z", f"{date_to}T23:59:59Z"
    try:
        gen = requests.get(
            f"{PERFORMANCE_API_URL}/api/client/statistics/all_sku_promo/orders/generate/json",
            params={"timeBounds.from": tb_from, "timeBounds.to": tb_to},
            headers=headers, timeout=30,
        )
        if gen.status_code != 200:
            out["error"] = f"generate вернул {gen.status_code}: {gen.text[:200]}"
            return out
        uuid = (gen.json() or {}).get("UUID")
        if not uuid:
            out["error"] = f"Нет UUID в ответе generate: {gen.text[:200]}"
            return out

        import time as _time
        state = None
        deadline = _time.time() + timeout_s
        while _time.time() < deadline:
            st_resp = requests.get(
                f"{PERFORMANCE_API_URL}/api/client/statistics/{uuid}",
                headers=headers, timeout=30,
            )
            if st_resp.status_code != 200:
                _time.sleep(3)
                continue
            st_data = st_resp.json() or {}
            state = st_data.get("state")
            if state == "OK":
                break
            if state == "ERROR":
                out["error"] = f"Ozon вернул ошибку формирования отчёта: {st_data.get('error')}"
                return out
            _time.sleep(3)
        else:
            out["error"] = f"Отчёт не сформировался за {timeout_s} сек (последний статус: {state})"
            return out

        rep = requests.get(
            f"{PERFORMANCE_API_URL}/api/client/statistics/report",
            params={"UUID": uuid}, headers=headers, timeout=60,
        )
        if rep.status_code != 200:
            out["error"] = f"report вернул {rep.status_code}: {rep.text[:200]}"
            return out

        rows = None
        try:
            parsed = rep.json()
            rows = parsed if isinstance(parsed, list) else (parsed.get("rows") or parsed.get("items"))
        except Exception:
            rows = None

        if rows is None:
            import csv, io
            text = rep.text
            delimiter = ";" if text.count(";") > text.count(",") else ","
            all_rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
            if len(all_rows) < 2:
                out["error"] = "Отчёт пуст или в неожиданном формате (не JSON и не похож на CSV)"
                return out
            header = [h.strip() for h in all_rows[0]]
            rows = [dict(zip(header, r)) for r in all_rows[1:]]

        if not rows:
            out["error"] = ""   # пусто — значит расхода по CPO за период реально нет, это не ошибка
            return out

        out["raw_sample"] = rows[:3]
        sample_keys = list(rows[0].keys()) if isinstance(rows[0], dict) else []
        # Подтверждено на реальном ответе API 02.08.2026: ключи date, order_id, order_number,
        # sku, adv_sku, vendor_code, name, quantity, price, sale_price, bid, abs_bid,
        # adv_money_spent. Берём "sku" (что реально куплено), НЕ "adv_sku" (что продвигалось —
        # может отличаться при кросс-показах), и "adv_money_spent" — это и есть расход, ₽.
        # Оставляем гибкий fallback на случай, если Ozon переименует поля в будущем.
        sku_key = next((k for k in sample_keys if str(k).strip().lower() == "sku"), None)
        expense_key = next((k for k in sample_keys if str(k).strip().lower() == "adv_money_spent"), None)
        if expense_key is None:
            expense_key = next(
                (k for k in sample_keys if any(t in str(k).lower() for t in ("expense", "расход", "spend", "spent"))),
                None,
            )
        date_key = next((k for k in sample_keys if str(k).strip().lower() == "date"), None)
        if sku_key is None or expense_key is None:
            out["error"] = f"Не распознала колонки SKU/расход в ответе API. Ключи в строке: {sample_keys}"
            return out

        totals: dict[str, float] = {}
        by_date: dict[str, dict[str, float]] = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            sku = str(row.get(sku_key) or "").strip()
            if not sku:
                continue
            try:
                val = float(str(row.get(expense_key) or "0").replace(",", "."))
            except Exception:
                val = 0.0
            totals[sku] = totals.get(sku, 0.0) + val
            if date_key:
                raw_d = str(row.get(date_key) or "").strip()
                iso_d = None
                # Подтверждённый формат из реального ответа — ДД.ММ.ГГГГ (как в файле).
                if raw_d and "." in raw_d:
                    parts = raw_d.split(".")
                    if len(parts) == 3 and len(parts[2]) == 4:
                        iso_d = f"{parts[2]}-{parts[1]}-{parts[0]}"
                elif len(raw_d) == 10 and raw_d[4] == "-":
                    iso_d = raw_d   # уже ISO (YYYY-MM-DD) — на случай другого формата в будущем
                if iso_d:
                    by_date.setdefault(iso_d, {})
                    by_date[iso_d][sku] = by_date[iso_d].get(sku, 0.0) + val
        out["by_sku"] = totals
        out["by_date"] = by_date
        out["total"] = sum(totals.values())
    except Exception as e:
        out["error"] = f"{e}"
    return out

def fetch_performance_ad_spend(
    perf_client_id: str, perf_client_secret: str, date_from: str, date_to: str,
    cpo_orders_file=None,
) -> dict:
    """
    Полный цикл: токен → (CPC: список SKU-кампаний → statistics/products/sku) +
    (CPO: файл пользователя ИЛИ best-effort async-отчёт all_sku_promo/orders) →
    объединённый расход по SKU. CPC и CPO обрабатываются НЕЗАВИСИМО — если один
    источник упал, второй всё равно применяется (см. cpc_ok/cpo_ok, используются
    в основном скрипте, чтобы не потерять и не задвоить деньги в "Расходах магазина").
    """
    out = {
        "by_sku": {}, "total": 0.0, "fetched": True, "error": None,
        "raw_campaigns_count": 0, "raw_campaigns_sample": [],
        "cpc_total": 0.0, "cpc_ok": False, "cpc_source": None, "cpc_campaigns_used": 0, "cpc_campaigns_failed": [], "cpc_by_sku": {}, "cpc_by_date": {}, "cpc_error_samples": [], "cpc_raw_sample": None,
        "cpo_total": 0.0, "cpo_ok": False, "cpo_source": None, "cpo_error": "", "cpo_raw_sample": None, "cpo_by_sku": {}, "cpo_by_date": {},
    }
    try:
        token = get_performance_token(perf_client_id, perf_client_secret)
    except Exception as e:
        out["error"] = f"Не удалось получить токен Performance API: {e}"
        return out

    # ── CPC: SKU-кампании ("Трафареты", оплата за клик) ──────────────────────
    # /statistics/products/sku (синхронный, "не расходует лимиты") принимает ТОЛЬКО
    # today/yesterday — задокументировано ("dateFrom не раньше предыдущего дня") и
    # подтверждено на реальном кабинете 03.08.2026 (400 "date range must contain only
    # today or yesterday" для всех кампаний разом при более раннем периоде). Если
    # запрошенный период уходит дальше вчера — используем асинхронный отчёт
    # (fetch_cpc_report_async), иначе — быстрый синхронный путь как раньше.
    cpc_by_sku: dict[str, float] = {}
    yesterday_iso = (date.today() - timedelta(days=1)).isoformat()
    _cpc_needs_async = date_from < yesterday_iso
    try:
        campaigns = fetch_performance_campaigns(token)
        out["raw_campaigns_count"] = len(campaigns)
        out["raw_campaigns_sample"] = [
            {
                "id": c.get("id"), "title": c.get("title"),
                "paymentType": c.get("paymentType"), "advObjectType": c.get("advObjectType"),
                "state": c.get("state"),
            }
            for c in campaigns[:30] if isinstance(c, dict)
        ]
        cpc_ids = [
            str(c.get("id")) for c in campaigns
            if isinstance(c, dict) and str(c.get("advObjectType") or "").upper() == "SKU"
        ]
        if cpc_ids and _cpc_needs_async:
            cpc_result = fetch_cpc_report_async(token, cpc_ids, date_from, date_to)
            cpc_by_sku = cpc_result.get("by_sku", {})
            out["cpc_by_date"] = cpc_result.get("by_date", {})
            out["cpc_source"] = "async"
            out["cpc_raw_sample"] = cpc_result.get("raw_sample")
            out["cpc_error_samples"] = cpc_result.get("chunks_failed", [])
            out["cpc_campaigns_used"] = len(cpc_ids) if cpc_by_sku else 0
            out["cpc_campaigns_failed"] = [] if cpc_by_sku else cpc_ids
            # ok, если хоть что-то получили, ИЛИ вообще не было ошибок по кускам периода
            out["cpc_ok"] = bool(cpc_by_sku) or not cpc_result.get("chunks_failed")
        elif cpc_ids:
            cpc_by_sku, cpc_by_date, failed, error_samples = fetch_performance_sku_expense(token, cpc_ids, date_from, date_to)
            out["cpc_by_date"] = cpc_by_date
            out["cpc_source"] = "sync"
            out["cpc_campaigns_used"] = len(cpc_ids) - len(failed)
            out["cpc_campaigns_failed"] = failed
            out["cpc_error_samples"] = error_samples
            out["cpc_ok"] = len(failed) < len(cpc_ids)   # хоть одна кампания отдала данные
        else:
            out["cpc_ok"] = True   # CPC-кампаний в кабинете просто нет — это не ошибка
    except Exception as e:
        out["error"] = f"Ошибка получения CPC-статистики: {e}"
    out["cpc_total"] = sum(cpc_by_sku.values())

    # ── CPO: "Оплата за заказ — все товары" ───────────────────────────────────
    cpo_by_sku: dict[str, float] = {}
    if cpo_orders_file is not None:
        cpo_by_sku, cpo_by_date, cpo_err = parse_all_sku_promo_orders_file(cpo_orders_file)
        out["cpo_by_date"] = cpo_by_date
        out["cpo_source"] = "file"
        out["cpo_error"] = cpo_err
        out["cpo_ok"] = not cpo_err
    else:
        cpo_result = fetch_all_sku_promo_orders_report(token, date_from, date_to)
        cpo_by_sku = cpo_result.get("by_sku", {})
        out["cpo_by_date"] = cpo_result.get("by_date", {})
        out["cpo_source"] = "api"
        out["cpo_error"] = cpo_result.get("error") or ""
        out["cpo_raw_sample"] = cpo_result.get("raw_sample")
        out["cpo_ok"] = not out["cpo_error"]
    out["cpo_total"] = sum(cpo_by_sku.values())
    out["cpc_by_sku"] = cpc_by_sku
    out["cpo_by_sku"] = cpo_by_sku

    # ── Объединяем ─────────────────────────────────────────────────────────
    combined: dict[str, float] = {}
    for d in (cpc_by_sku, cpo_by_sku):
        for k, v in d.items():
            combined[k] = combined.get(k, 0.0) + v
    out["by_sku"] = combined
    out["total"] = sum(combined.values())

    if not out["cpc_ok"] and not out["cpo_ok"] and not out["error"]:
        out["error"] = "Не удалось получить данные ни по CPC, ни по CPO — см. диагностику (Tab5)."
    return out

# ── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("⬡ Ozon P&L")
    st.caption("Дашборд юнит-экономики")
    st.caption(f"⚙️ Версия от {APP_BUILD_VERSION}")

    st.divider()
    st.subheader("🔑 API-доступ")
    client_id = st.text_input(
        "Client-ID", value=_secret("OZON_CLIENT_ID"), placeholder="123456",
        help="Seller API: seller.ozon.ru → Настройки → API-ключи. Чтобы не вводить каждый раз — "
             "добавь ключи в Secrets приложения (⋮ внизу справа → Manage app → Settings → Secrets): "
             "OZON_CLIENT_ID, OZON_API_KEY, OZON_PERF_CLIENT_ID, OZON_PERF_CLIENT_SECRET — тогда "
             "поля заполнятся сами."
    )
    api_key = st.text_input("API-Key", value=_secret("OZON_API_KEY"), type="password", placeholder="xxxx-xxxx-xxxx")

    with st.expander("🎯 Реклама по артикулам (Performance API)"):
        perf_client_id = st.text_input(
            "Performance Client-ID", value=_secret("OZON_PERF_CLIENT_ID"),
            placeholder="123@advertising.performance.ozon.ru",
            help="Отдельная пара ключей от Seller API выше — Настройки → API-ключи → сервисный "
                 "аккаунт с доступом к Performance API. Даёт расход на рекламу (CPC/CPO) по каждому "
                 "артикулу вместо одной суммы на весь магазин."
        )
        perf_client_secret = st.text_input(
            "Performance Client-Secret", value=_secret("OZON_PERF_CLIENT_SECRET"),
            type="password", placeholder="xxxx-xxxx-xxxx",
        )
        use_perf_ads = st.checkbox(
            "Учитывать в прибыли по артикулам", value=True,
            help="Вкл: сумма из Performance API ЗАМЕНЯЕТ статью «Реклама» (CPC/CPO) в «Расходах "
                 "магазина» и распределяется по артикулам — без задвоения. Выкл: как раньше, "
                 "одной суммой на магазин."
        )
        cpo_orders_file = st.file_uploader(
            "Отчёт «Оплата за заказ» (xlsx/csv)", type=["xlsx", "csv"],
            help="Нужен, только если авто-запрос CPO не сработает («Оплата за заказ» отдаёт расход "
                 "по SKU не через тот же метод, что «Оплата за клик», и авто-запрос экспериментальный). "
                 "Скачивается в ЛК Ozon: Статистика → Оплата за заказ, все товары → Отчёт по заказам. "
                 "Если файл загружен — используется вместо API."
        )

    st.divider()
    st.subheader("📅 Период")
    mode = st.radio("Режим", ["Текущий месяц", "Прошлый месяц", "Произвольные даты"])

    today = date.today()
    if mode == "Текущий месяц":
        d_from = today.replace(day=1)
        d_to = today
        use_transactions = True
    elif mode == "Прошлый месяц":
        first_this = today.replace(day=1)
        d_to = first_this - timedelta(days=1)
        d_from = d_to.replace(day=1)
        use_transactions = True
    else:
        d_from = st.date_input("Дата от", value=today.replace(day=1))
        d_to = st.date_input("Дата до", value=today)
        use_transactions = not (d_from.day == 1 and d_to == date(d_from.year, d_from.month,
            calendar.monthrange(d_from.year, d_from.month)[1]))

    st.caption(f"{'Транзакции' if use_transactions else 'Отчёт реализации'}: {d_from} — {d_to}")

    st.divider()
    st.subheader("💰 Себестоимость")

    cost_file = st.file_uploader(
        "Загрузи файл Excel/CSV: колонки «артикул» и «себестоимость»",
        type=["xlsx", "csv"],
        help="При изменении цен — просто загрузи новый файл"
    )

    if "cost_history" not in st.session_state:
        st.session_state.cost_history = []

    cost_map = {}

    if cost_file is not None:
        try:
            if cost_file.name.endswith(".csv"):
                cost_df = pd.read_csv(cost_file)
            else:
                cost_df = pd.read_excel(cost_file)

            cost_df.columns = [c.strip().lower() for c in cost_df.columns]
            art_col = next((c for c in cost_df.columns if "артикул" in c or "article" in c or "sku" in c), None)
            cost_col = next((c for c in cost_df.columns if "себест" in c or "cost" in c or "закуп" in c or "цена" in c), None)

            if art_col and cost_col:
                cost_df = cost_df.dropna(subset=[art_col])
                cost_df[art_col] = cost_df[art_col].astype(str).str.strip()
                cost_df = cost_df[cost_df[art_col].str.len() > 0]
                cost_df[cost_col] = (
                    cost_df[cost_col]
                    .astype(str)
                    .str.replace('\xa0', '', regex=False)   # неразрывный пробел (1 043,00)
                    .str.replace(' ', '', regex=False)  # тонкий пробел
                    .str.replace(' ', '', regex=False)       # обычный пробел
                    .str.replace(',', '.', regex=False)      # запятая → точка
                    .pipe(pd.to_numeric, errors='coerce')
                )
                cost_map = dict(zip(cost_df[art_col], cost_df[cost_col].fillna(0)))

                already = any(h["filename"] == cost_file.name and
                              h["date"] == date.today().strftime("%d.%m.%Y")
                              for h in st.session_state.cost_history)
                if not already:
                    st.session_state.cost_history.append({
                        "date": date.today().strftime("%d.%m.%Y"),
                        "filename": cost_file.name,
                        "count": len(cost_map),
                        "data": cost_map.copy(),
                    })
                st.success(f"✅ Загружено {len(cost_map)} артикулов")
            else:
                st.error("Не нашла колонки. Нужны: «артикул» и «себестоимость»")
        except Exception as e:
            st.error(f"Ошибка: {e}")

    if not cost_map and st.session_state.cost_history:
        last = st.session_state.cost_history[-1]
        cost_map = last["data"]
        st.info(f"Используется: {last['filename']} от {last['date']}")

    if st.session_state.cost_history:
        with st.expander(f"📜 История загрузок ({len(st.session_state.cost_history)})"):
            for h in reversed(st.session_state.cost_history):
                st.caption(f"📅 {h['date']} — {h['filename']} ({h['count']} артикулов)")

    load_btn = st.button("🔄 Загрузить данные", type="primary", use_container_width=True)
    demo_btn = st.button("🎲 Демо-данные", use_container_width=True)
    debug_btn = st.button("🔍 Показать сырые данные API", use_container_width=True)

# ── Demo data ────────────────────────────────────────────────────────────────
def make_demo() -> pd.DataFrame:
    import random
    random.seed(42)
    skus = [
        ("11860",   "pjur Woman Nude 100мл",      1204, 3251, 45, 289),
        ("6033487", "Фаллоим. UTOPIA S 13.5см",   3100, 7400, 40, 322),
        ("O17182",  "Orgie Cocktail Тропик 50мл",  1100, 2769, 39, 195),
        ("6022351", "MixGliss Nature 150мл",        1800, 3410, 39, 238),
        ("3003Un",  "Unilatex Natural 3шт",           75,  299, 29,  42),
        ("3017Un",  "Unilatex Multifruits 3шт",       75,  309, 29,  42),
        ("541438",  "Erotist VIVID ACT 30мл",         320, 1478, 45, 110),
        ("6033081", "LoveToLove SUNRISE",            3900, 7965, 40, 276),
        ("602211",  "MixGliss SUN 50мл",              900, 3733, 39, 195),
        ("51652",   "Orgie Orgasm Drops Vibe!",      1200, 3139, 39, 168),
    ]
    rows = []
    for art, name, cost, price, comm_pct, log in skus:
        qty = random.randint(3, 45)
        qret = max(0, random.randint(0, 2))
        rev = price * qty
        comm = (comm_pct / 100) * rev
        acq = ACQUIRING * rev
        tx = TAX * rev
        logi = log * qty
        rows.append({
            "article": art,
            "name": name,
            "qty": qty,
            "qty_ret": qret,
            "revenue": rev,
            "return_sum": -(price * qret),
            "commission": -comm,
            "logistics": -logi,
            "cost_price": cost,
            "cost_total": cost * qty,
            "acquiring": acq,
            "tax": tx,
            "profit": rev - cost * qty - comm - acq - tx - logi,
            "other": 0,
        })
    df = pd.DataFrame(rows)
    df["margin_pct"] = (df["profit"] / df["revenue"].replace(0, float("nan"))) * 100
    return df

# ── State ────────────────────────────────────────────────────────────────────
if "df" not in st.session_state:
    st.session_state.df = None
if "is_demo" not in st.session_state:
    st.session_state.is_demo = False
if "raw_ops" not in st.session_state:
    st.session_state.raw_ops = []
if "sku_map_cache" not in st.session_state:
    st.session_state.sku_map_cache = {}
if "store_costs" not in st.session_state:
    st.session_state.store_costs = {}
if "loaded_period" not in st.session_state:
    st.session_state.loaded_period = None  # (d_from, d_to) последней успешной загрузки
if "stocks_data" not in st.session_state:
    st.session_state.stocks_data = []
if "perf_ad_spend" not in st.session_state:
    st.session_state.perf_ad_spend = {}   # {sku: расход ₽} из Performance API
if "perf_ad_meta" not in st.session_state:
    st.session_state.perf_ad_meta = {
        "total": 0.0, "fetched": False, "error": None,
        "raw_campaigns_count": 0, "raw_campaigns_sample": [],
        "cpc_total": 0.0, "cpc_ok": False, "cpc_source": None, "cpc_campaigns_used": 0, "cpc_campaigns_failed": [], "cpc_by_sku": {}, "cpc_by_date": {}, "cpc_error_samples": [], "cpc_raw_sample": None,
        "cpo_total": 0.0, "cpo_ok": False, "cpo_source": None, "cpo_error": "", "cpo_raw_sample": None, "cpo_by_sku": {}, "cpo_by_date": {},
    }

if demo_btn:
    st.session_state.df = make_demo()
    st.session_state.is_demo = True

if debug_btn:
    if not client_id or not api_key:
        st.error("Введи Client-ID и API-Key в боковом меню")
    else:
        st.subheader("🔍 Сырые данные API (для диагностики)")
        today_str = date.today().strftime("%Y-%m-%d")
        prev_month = date.today().replace(day=1) - timedelta(days=1)

        with st.expander("📄 /v2/finance/realization — прошлый месяц (первые 2 строки)", expanded=True):
            with st.spinner("Запрос..."):
                raw_real = api_post(
                    "/v2/finance/realization",
                    {"year": prev_month.year, "month": prev_month.month},
                    client_id, api_key,
                )
            result = (raw_real.get("result") or {}) if raw_real else {}
            rows = (result.get("rows") or [])[:2]
            if rows:
                import json
                st.code(json.dumps(rows, ensure_ascii=False, indent=2), language="json")
            else:
                st.warning("Нет данных (отчёт формируется в начале следующего месяца)")
                st.json(raw_real or {})

        with st.expander("📄 /v1/finance/accrual/by-day — вчера (первые 2 начисления)", expanded=True):
            yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
            with st.spinner("Запрос..."):
                raw_tx = api_post(
                    "/v1/finance/accrual/by-day",
                    {"date": yesterday, "last_id": ""},
                    client_id, api_key,
                )
            accruals = (raw_tx.get("accruals") or [])[:2] if raw_tx else []
            if accruals:
                import json
                st.code(json.dumps(accruals, ensure_ascii=False, indent=2), language="json")
            else:
                st.warning("Нет начислений за вчера")
                st.json(raw_tx or {})

        st.info("Скопируй содержимое блоков выше и пришли — найдём поле эквайринга.")

if load_btn:
    if not client_id or not api_key:
        st.error("Введи Client-ID и API-Key в боковом меню")
    else:
        # Сбрасываем старые данные ДО запроса — чтобы не показывать чужой период при ошибке
        st.session_state.df = None
        st.session_state.loaded_period = None
        st.session_state.store_costs = {}
        st.session_state.perf_ad_spend = {}
        st.session_state.perf_ad_meta = {
            "total": 0.0, "fetched": False, "error": None,
            "raw_campaigns_count": 0, "raw_campaigns_sample": [],
            "cpc_total": 0.0, "cpc_ok": False, "cpc_campaigns_used": 0, "cpc_campaigns_failed": [], "cpc_by_sku": {}, "cpc_by_date": {}, "cpc_error_samples": [],
            "cpo_total": 0.0, "cpo_ok": False, "cpo_source": None, "cpo_error": "", "cpo_raw_sample": None, "cpo_by_sku": {}, "cpo_by_date": {},
        }
        with st.spinner("Загружаем данные из Ozon API..."):
            try:
                if use_transactions:
                    ops = fetch_transactions(client_id, api_key,
                                             d_from.strftime("%Y-%m-%d"),
                                             d_to.strftime("%Y-%m-%d"))
                    raw_df = transactions_to_df(ops)
                else:
                    rows_api = fetch_realization(client_id, api_key, d_from.year, d_from.month)
                    raw_df = realization_to_df(rows_api)

                if raw_df.empty:
                    st.warning("Нет данных за выбранный период. Отчёт реализации формируется Ozon в начале следующего месяца.")
                else:
                    # Загружаем справочник типов начислений (type_id → название)
                    with st.spinner("Загружаем справочник типов начислений..."):
                        api_type_names = fetch_accrual_types(client_id, api_key)
                        if api_type_names:
                            TYPE_NAMES.update(api_type_names)

                    # Загружаем справочник sku → offer_id
                    # Сначала быстрый путь: turnover/stocks (Product read-only)
                    with st.spinner("Загружаем справочник артикулов..."):
                        _unique_skus = tuple(sorted(raw_df["sku"].unique().tolist())) if "sku" in raw_df.columns else ()
                        _offer_id_map = fetch_offer_ids_by_sku(client_id, api_key, _unique_skus)
                    if _offer_id_map:
                        raw_df["article"] = raw_df["sku"].map(_offer_id_map).fillna(raw_df["article"])
                        raw_df["article"] = raw_df["article"].astype(str)
                        sku_map = _offer_id_map
                    else:
                        # Запасной путь: FBO/FBS list
                        with st.spinner("Запрашиваем артикулы через FBO/FBS..."):
                            sku_map = fetch_sku_map(client_id, api_key)
                        if sku_map:
                            name_map = fetch_name_map(client_id, api_key)
                            raw_df = apply_sku_map(raw_df, sku_map, name_map)
                        else:
                            st.warning("Не удалось получить артикулы: проверь права API-ключа (нужен Product read-only или FBO/FBS)")

                    if use_transactions:
                        # Транзакционный режим — логистика уже в raw_df
                        st.session_state.raw_ops = ops
                        st.session_state.sku_map_cache = sku_map if sku_map else {}
                        st.session_state.store_costs = collect_store_costs(ops)
                    else:
                        # Реализация: транзакции для логистики + расходов магазина
                        # Запрашиваем ДО enrich_with_cost, чтобы логистика попала в расчёт прибыли
                        with st.spinner("Загружаем логистику и расходы магазина..."):
                            try:
                                ops_for_costs = fetch_transactions(
                                    client_id, api_key,
                                    d_from.strftime("%Y-%m-%d"),
                                    d_to.strftime("%Y-%m-%d"),
                                )
                                st.session_state.store_costs = collect_store_costs(ops_for_costs)
                                st.session_state.raw_ops = ops_for_costs
                                st.session_state.sku_map_cache = sku_map if sku_map else {}

                                # Суммируем логистику по артикулам из транзакций
                                _logi: dict[str, float] = {}
                                for _a in ops_for_costs:
                                    _p = _a.get("posting") or {}
                                    for _pr in (_p.get("products") or []):
                                        if not isinstance(_pr, dict):
                                            continue
                                        _sku = str(_pr.get("sku") or "")
                                        if not _sku:
                                            continue
                                        _oid = (sku_map or {}).get(_sku, _sku)
                                        _d = _pr.get("delivery") or {}
                                        _v = float(((_d.get("total_accrued") or {}).get("amount") or 0))
                                        if _v:
                                            _logi[_oid] = _logi.get(_oid, 0.0) + _v
                                # Обновляем logistics в raw_df ДО enrich_with_cost
                                if _logi:
                                    raw_df["logistics"] = raw_df["article"].map(_logi).fillna(0.0)
                            except Exception:
                                pass  # расходы магазина необязательны

                    # Реклама CPC/CPO по артикулам (Performance API) — опционально,
                    # отдельная авторизация от Seller API. Не блокирует основную загрузку при ошибке.
                    if perf_client_id and perf_client_secret:
                        with st.spinner("Загружаем расход на рекламу (Performance API)..."):
                            perf_result = fetch_performance_ad_spend(
                                perf_client_id, perf_client_secret,
                                d_from.strftime("%Y-%m-%d"), d_to.strftime("%Y-%m-%d"),
                                cpo_orders_file=cpo_orders_file,
                            )
                        st.session_state.perf_ad_spend = perf_result["by_sku"]
                        st.session_state.perf_ad_meta = perf_result

                        if perf_result["error"]:
                            st.warning(
                                f"⚠️ Performance API: {perf_result['error']} — реклама по артикулам "
                                f"не будет учтена в этой загрузке, «Расходы магазина» остаются как раньше "
                                f"(одна сумма CPC/CPO на весь магазин из Seller API)."
                            )
                        else:
                            _cpc_src_label = {"sync": "быстрый, today/yesterday", "async": "отчёт по периоду"}.get(
                                perf_result.get("cpc_source"), perf_result.get("cpc_source")
                            )
                            st.success(
                                f"🎯 Performance API: CPC {r(perf_result['cpc_total'])} "
                                f"({'ок, источник: ' + str(_cpc_src_label) if perf_result['cpc_ok'] else 'не удалось'}) + "
                                f"CPO {r(perf_result['cpo_total'])} "
                                f"({'ок, источник: ' + str(perf_result['cpo_source']) if perf_result['cpo_ok'] else 'не удалось'})"
                            )
                        if not perf_result["cpc_ok"]:
                            if perf_result.get("cpc_source") == "async":
                                st.warning(
                                    "⚠️ CPC (оплата за клик): не удалось получить исторический отчёт за период — "
                                    "эта часть расхода останется в «Расходах магазина» одной суммой."
                                )
                            else:
                                st.warning(
                                    f"⚠️ CPC (оплата за клик): не удалось получить статистику "
                                    f"({len(perf_result['cpc_campaigns_failed'])} кампаний с ошибкой) — "
                                    f"эта часть расхода останется в «Расходах магазина» одной суммой."
                                )
                            if perf_result.get("cpc_error_samples"):
                                with st.expander("🔍 Реальный текст ошибки от Performance API", expanded=True):
                                    st.dataframe(
                                        pd.DataFrame(perf_result["cpc_error_samples"]),
                                        use_container_width=True, hide_index=True,
                                    )
                        elif perf_result["cpc_campaigns_failed"] and perf_result.get("cpc_source") != "async":
                            st.warning(
                                f"⚠️ CPC: {len(perf_result['cpc_campaigns_failed'])} из "
                                f"{perf_result['cpc_campaigns_used'] + len(perf_result['cpc_campaigns_failed'])} "
                                f"кампаний реально проблемные (изолированы, остальные подтянулись нормально) — "
                                f"расход по ним не учтён."
                            )
                            if perf_result.get("cpc_error_samples"):
                                with st.expander("🔍 Какие кампании и почему не подтянулись"):
                                    st.dataframe(
                                        pd.DataFrame(perf_result["cpc_error_samples"]),
                                        use_container_width=True, hide_index=True,
                                    )
                        elif perf_result.get("cpc_source") == "async" and perf_result.get("cpc_error_samples"):
                            with st.expander(f"🔍 Часть периода не подтянулась ({len(perf_result['cpc_error_samples'])} кусков) — детали"):
                                st.dataframe(
                                    pd.DataFrame(perf_result["cpc_error_samples"]),
                                    use_container_width=True, hide_index=True,
                                )
                        if perf_result["cpo_error"]:
                            st.warning(f"⚠️ CPO (оплата за заказ): {perf_result['cpo_error']}")

                        if perf_result.get("raw_campaigns_sample") and (perf_result["error"] or not perf_result["cpc_ok"]):
                            with st.expander(
                                f"🔍 Сырые кампании из Performance API "
                                f"(всего {perf_result.get('raw_campaigns_count', 0)}, показаны первые "
                                f"{len(perf_result['raw_campaigns_sample'])})",
                                expanded=True,
                            ):
                                st.dataframe(
                                    pd.DataFrame(perf_result["raw_campaigns_sample"]),
                                    use_container_width=True, hide_index=True,
                                )
                        if perf_result.get("cpc_raw_sample"):
                            with st.expander(
                                "🔍 Сырой пример строк CPC-отчёта за период (для диагностики — схема полей "
                                "не задокументирована Ozon, парсинг может понадобиться поправить по этому примеру)",
                                expanded=not perf_result["cpc_ok"],
                            ):
                                st.json(perf_result["cpc_raw_sample"])
                        if perf_result.get("cpo_raw_sample"):
                            with st.expander("🔍 Сырой пример строк CPO-отчёта (для диагностики)", expanded=bool(perf_result["cpo_error"])):
                                st.json(perf_result["cpo_raw_sample"])

                    _perf_cpc_map = st.session_state.perf_ad_meta.get("cpc_by_sku", {}) if use_perf_ads else {}
                    _perf_cpo_map = st.session_state.perf_ad_meta.get("cpo_by_sku", {}) if use_perf_ads else {}
                    # enrich_with_cost вызывается ПОСЛЕ обновления логистики
                    st.session_state.df = enrich_with_cost(
                        raw_df, cost_map, cpc_ad_map=_perf_cpc_map, cpo_ad_map=_perf_cpo_map,
                    )
                    st.session_state._cost_map_debug = cost_map  # для диагностики в Tab5
                    st.session_state.is_demo = False
                    st.session_state.loaded_period = (d_from, d_to)
                    with st.spinner("Загружаем остатки и оборачиваемость..."):
                        skus_tuple = tuple(sorted(sku_map.keys())) if sku_map else ()
                        st.session_state.stocks_data = fetch_stocks(client_id, api_key, skus_tuple)
            except Exception as e:
                st.error(f"Ошибка загрузки: {e}")

# ── Main ─────────────────────────────────────────────────────────────────────
df = st.session_state.df

if df is None:
    st.title("📊 Ozon P&L Dashboard")
    st.info("👈 Введи API-ключи и нажми **Загрузить данные** — или попробуй **Демо-данные**")
    with st.expander("Как получить API-ключи?"):
        st.markdown("""
1. Зайди в [seller.ozon.ru](https://seller.ozon.ru)
2. **Настройки → API-ключи**
3. Нажми «Создать ключ», выбери роль **Финансы** (или Администратор)
4. Скопируй **Client-ID** и **API-Key** → вставь в меню слева
        """)
    st.stop()

if st.session_state.is_demo:
    st.warning("⚠️ Демо-данные. Введи API-ключи для реальных цифр.", icon="🎲")

# ── Заголовок ─────────────────────────────────────────────────────────────────
loaded_period = st.session_state.get("loaded_period")
if loaded_period and not st.session_state.is_demo:
    loaded_from, loaded_to = loaded_period
    if loaded_from != d_from or loaded_to != d_to:
        st.warning(
            f"⚠️ Данные загружены за **{loaded_from.strftime('%d.%m.%Y')} — {loaded_to.strftime('%d.%m.%Y')}**, "
            f"а выбран период **{d_from.strftime('%d.%m.%Y')} — {d_to.strftime('%d.%m.%Y')}**. "
            "Нажми **Загрузить данные** для обновления."
        )
    days = (loaded_to - loaded_from).days + 1
    st.title(f"📊 P&L: {loaded_from.strftime('%d.%m')} — {loaded_to.strftime('%d.%m.%Y')}")
else:
    days = (d_to - d_from).days + 1
    st.title(f"📊 P&L: {d_from.strftime('%d.%m')} — {d_to.strftime('%d.%m.%Y')}")

# ── KPI ───────────────────────────────────────────────────────────────────────
total_rev  = df["revenue"].sum()
total_partner = df["partner_programs"].sum() if "partner_programs" in df.columns else 0.0
total_bonus    = df["bonus_points"].sum() if "bonus_points" in df.columns else 0.0
total_income   = df["total_income"].sum() if "total_income" in df.columns else total_rev
total_prof = df["profit"].sum()
total_comm = df["commission"].abs().sum()
total_log  = df["logistics"].abs().sum() if "logistics" in df.columns else 0
total_cost = df["cost_total"].sum() if "cost_total" in df.columns else 0
total_qty  = df["qty"].sum()

# Расходы магазина нужны ДО метрик — чтобы показать реальную прибыль и правильную маржу
_sc_kpi: dict = st.session_state.get("store_costs", {})
_perf_meta_kpi: dict = st.session_state.get("perf_ad_meta", {}) or {}
# Реклама CPC (type_id 41) и CPO (type_id 54) заменяются на Performance API НЕЗАВИСИМО
# друг от друга — только тот из двух, который реально успешно получен (cpc_ok/cpo_ok).
# Так расход не теряется, если один источник не сработал, а другой — сработал.
_perf_replaces_cpc = bool(use_perf_ads and _perf_meta_kpi.get("fetched") and _perf_meta_kpi.get("cpc_ok"))
_perf_replaces_cpo = bool(use_perf_ads and _perf_meta_kpi.get("fetched") and _perf_meta_kpi.get("cpo_ok"))
_perf_replaces_store_ads = _perf_replaces_cpc or _perf_replaces_cpo   # хоть что-то заменено
_excluded_store_type_ids = set()
if _perf_replaces_cpc:
    _excluded_store_type_ids.add(PERFORMANCE_CPC_STORE_TYPE_ID)
if _perf_replaces_cpo:
    _excluded_store_type_ids.add(PERFORMANCE_CPO_STORE_TYPE_ID)
_sc_kpi_effective = {tid: amt for tid, amt in _sc_kpi.items() if tid not in _excluded_store_type_ids}
_sc_total_kpi = sum(v for v in _sc_kpi_effective.values() if v < 0) if _sc_kpi_effective else 0

# SKU, у которых был расход в Performance API, но не оказалось строки в df (например,
# показы/клики были, а продаж в выбранном периоде — нет). Эти деньги реальны и уже списаны
# Ozon — чтобы не потерять их из P&L, довычитаем отдельно, вне таблицы по артикулам.
_perf_unmatched = 0.0
if _perf_replaces_store_ads:
    _perf_matched_total = df["ads_perf"].sum() if "ads_perf" in df.columns else 0.0
    _perf_unmatched = max(0.0, float(_perf_meta_kpi.get("total", 0.0)) - float(_perf_matched_total))

total_prof_adj = total_prof + _sc_total_kpi - _perf_unmatched   # реальная прибыль = прибыль по артикулам − расходы магазина − несопоставленная реклама
# Маржа считается от полного дохода (Выручка + Программы партнёров + Баллы за скидки) —
# это соответствует построчному margin_pct в enrich_with_cost().
total_margin = (total_prof_adj / total_income * 100) if total_income else 0

comm_pct = (total_comm / total_rev * 100) if total_rev else 0
log_pct  = (total_log  / total_rev * 100) if total_rev else 0
cost_pct = (total_cost / total_rev * 100) if total_rev else 0

total_tax = df["tax"].sum() if "tax" in df.columns else 0.0

c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
c1.metric("Выручка", r(total_rev), f"{total_rev/days:,.0f} ₽/день".replace(",", " "))
if total_partner or total_bonus:
    c1.caption(f"из них: Программы {r(total_partner)}, Баллы {r(total_bonus)}")
c2.metric("Реальная прибыль", r(total_prof_adj), f"{total_prof_adj/days:,.0f} ₽/день".replace(",", " "),
          delta_color="normal" if total_prof_adj >= 0 else "inverse")
c3.metric("Маржа", ru_pct(total_margin), delta_color="normal" if total_margin >= 5 else "inverse")
c4.metric("Продано", f"{int(total_qty)} шт", f"{total_qty/days:.1f}".replace(".", ",") + " шт/день")
with c5:
    st.metric("Себестоимость", r(total_cost))
    st.caption(f"({ru_pct(cost_pct)} от выручки)")
with c6:
    st.metric("Комиссия", r(total_comm))
    st.caption(f"({ru_pct(comm_pct)} от выручки)")
with c7:
    st.metric("Логистика", r(total_log))
    st.caption(f"({ru_pct(log_pct)} от выручки)")

# ── Расходы магазина (non_item_fee) ─────────────────────────────────────────
store_costs: dict = _sc_kpi_effective   # уже загружено выше для KPI (без 41/54, если заменены Performance API)
if store_costs or total_tax or _perf_unmatched:
    store_total = _sc_total_kpi   # уже вычислено выше
    # total_prof_adj уже вычислен выше

    with st.expander(
        f"🏪 Расходы магазина: {r(abs(store_total) + _perf_unmatched)} + Налог {r(total_tax)} "
        f"(налог не входит в таблицу по артикулам отдельной строкой расхода магазина — "
        f"он уже вычтен из «Прибыли по артикулам»; показан здесь просто для сводки)",
        expanded=False,
    ):
        if _perf_replaces_store_ads:
            _replaced_names = []
            if _perf_replaces_cpc:
                _replaced_names.append("CPC (оплата за клик, type 41)")
            if _perf_replaces_cpo:
                _replaced_names.append("CPO (оплата за заказ, type 54)")
            st.info(
                f"🎯 Реклама {' и '.join(_replaced_names)} теперь считается по Performance API "
                f"и распределена по артикулам (колонка «Реклама CPC/CPO» в таблице ниже) — "
                f"вместо одной суммы здесь, чтобы не задвоить расход."
                + ("" if _perf_replaces_cpc and _perf_replaces_cpo else
                   " Вторая часть (не заменённая) пока остаётся здесь одной суммой, как раньше.")
            )
        sc_rows = []
        for tid, amt in sorted(store_costs.items(), key=lambda x: x[1]):
            if amt == 0:
                continue
            sc_rows.append({
                "Статья": TYPE_NAMES.get(tid, f"type_{tid}"),
                "Группа": STORE_COST_GROUPS.get(tid, "Прочее"),
                "Сумма": amt,
            })
        # Налог УСН — отдельной строкой в этой же сводке (по запросу пользователя, 02.08.2026).
        if total_tax:
            sc_rows.append({
                "Статья": "Налог УСН 7% (Выручка + Программы партнёров)",
                "Группа": "Налоги",
                "Сумма": -abs(total_tax),
            })
        if _perf_unmatched:
            sc_rows.append({
                "Статья": "Реклама на SKU без продаж в периоде (Performance API)",
                "Группа": "Реклама",
                "Сумма": -abs(_perf_unmatched),
            })
        if sc_rows:
            sc_df = pd.DataFrame(sc_rows)
            by_group = sc_df.groupby("Группа")["Сумма"].sum().reset_index().sort_values("Сумма")
            col_sc1, col_sc2 = st.columns([1, 2])
            with col_sc1:
                st.markdown("**По группам:**")
                for _, row in by_group.iterrows():
                    st.write(f"{row['Группа']}: **{r(abs(row['Сумма']))}**")
            with col_sc2:
                st.markdown("**Детализация:**")
                st.dataframe(
                    sc_df.style.format({"Сумма": lambda v: ru_float(v, 2)}),
                    use_container_width=True, hide_index=True,
                )
        sa1, sa2, sa3, sa4 = st.columns(4)
        sa1.metric("Прибыль по артикулам", r(total_prof))
        sa2.metric("Расходы магазина", r(abs(store_total) + _perf_unmatched))
        sa3.metric("Налог (справочно)", r(total_tax))
        sa4.metric("Реальная прибыль", r(total_prof_adj),
                   delta_color="normal" if total_prof_adj >= 0 else "inverse")

        if _perf_replaces_cpc:
            _old_cpc = abs(_sc_kpi.get(PERFORMANCE_CPC_STORE_TYPE_ID, 0.0))
            _new_cpc = abs(float(_perf_meta_kpi.get("cpc_total", 0.0)))
            st.caption(
                f"🔍 Сверка CPC — Seller API (type 41, как считалось раньше) = **{r(_old_cpc)}**; "
                f"Performance API (сумма по SKU, из чего сейчас считается P&L) = **{r(_new_cpc)}**; "
                f"разница = **{r(abs(_old_cpc - _new_cpc))}**."
            )
            if _perf_meta_kpi.get("cpc_campaigns_failed"):
                st.warning(
                    f"⚠️ {len(_perf_meta_kpi['cpc_campaigns_failed'])} CPC-кампаний не ответили "
                    f"при этой загрузке — сумма CPC выше может быть занижена."
                )
        if _perf_replaces_cpo:
            _old_cpo = abs(_sc_kpi.get(PERFORMANCE_CPO_STORE_TYPE_ID, 0.0))
            _new_cpo = abs(float(_perf_meta_kpi.get("cpo_total", 0.0)))
            st.caption(
                f"🔍 Сверка CPO — Seller API (type 54, как считалось раньше) = **{r(_old_cpo)}**; "
                f"Performance API (источник: {_perf_meta_kpi.get('cpo_source')}, из чего сейчас "
                f"считается P&L) = **{r(_new_cpo)}**; разница = **{r(abs(_old_cpo - _new_cpo))}**."
            )
        if _perf_replaces_cpc or _perf_replaces_cpo:
            st.caption(
                "Расхождение возможно из-за разницы в моделях учёта дат/списаний между Seller "
                "API и Performance API — если оно большое, стоит уточнить у поддержки Ozon, а "
                "не считать ошибкой кода."
            )

st.divider()

# ── Страницы (левая навигация, st.navigation) ────────────────────────────────
# Раньше здесь были st.tabs() с 5 вкладками (включая Диаграммы и Калькулятор — оба
# убраны по просьбе пользователя 02.08.2026). Функции ниже читают общие переменные
# (df, total_rev, days, client_id, api_key, sku_map_cache, raw_ops и т.д.) как обычные
# глобальные — они уже вычислены выше по скрипту к моменту вызова pg.run() в самом
# конце файла, поэтому доступны без явной передачи параметров.
def page_articles():
    show_cols = ["article", "name", "qty", "qty_ret", "revenue", "partner_programs", "bonus_points", "cost_total",
                 "commission", "acquiring", "tax", "logistics", "promo", "ads_cpc", "ads_cpo", "installment", "other_costs", "profit", "margin_pct"]
    available = [c for c in show_cols if c in df.columns]
    display_df = df[available].copy()

    rename = {
        "article": "Артикул",
        "name": "Товар",
        "qty": "Продано",
        "qty_ret": "Возвращено",
        "revenue": "Выручка",
        "partner_programs": "Программы партнёров",
        "bonus_points": "Баллы за скидки",
        "cost_total": "Себестоимость",
        "commission": "Комиссия",
        "acquiring": "Эквайринг",
        "tax": "Налог",
        "logistics": "Логистика",
        "promo":       "Реклама",
        "ads_cpc":     "Реклама CPC",
        "ads_cpo":     "Реклама CPO",
        "installment": "Рассрочка",
        "other_costs": "Прочие расходы",
        "profit": "Прибыль",
        "margin_pct": "Маржа %",
    }
    display_df = display_df.rename(columns=rename).sort_values("Прибыль", ascending=False)

    # Форматирование с пробелом как разделителем тысяч
    def _fmt_rub(v):
        if pd.isna(v): return "—"
        return ru_rub(v, 0)

    rub_cols = ["Выручка", "Программы партнёров", "Баллы за скидки", "Себестоимость", "Комиссия", "Эквайринг", "Налог", "Логистика", "Реклама", "Реклама CPC", "Реклама CPO", "Рассрочка", "Прочие расходы", "Прибыль"]
    fmt_dict = {c: _fmt_rub for c in rub_cols if c in display_df.columns}
    if "Маржа %" in display_df.columns:
        fmt_dict["Маржа %"] = ru_pct

    def color_margin(val):
        if pd.isna(val):
            return ""
        if val >= 10:
            return "color: #3DD68C; font-weight: bold"
        if val >= 5:
            return "color: #F0A93D"
        return "color: #F05B5B"

    def color_profit(val):
        if pd.isna(val):
            return ""
        return "color: #3DD68C" if val >= 0 else "color: #F05B5B"

    styled = (
        display_df.style
        .format(fmt_dict, na_rep="—")
        .map(color_margin, subset=["Маржа %"] if "Маржа %" in display_df.columns else [])
        .map(color_profit, subset=["Прибыль"] if "Прибыль" in display_df.columns else [])
    )
    st.dataframe(styled, use_container_width=True, height=420)

    # Итого
    st.markdown("**Итого по всем артикулам:**")
    _total_promo = df["promo"].abs().sum() if "promo" in df.columns else 0
    _total_acq   = df["acquiring"].abs().sum() if "acquiring" in df.columns else 0
    _total_inst  = df["installment"].abs().sum() if "installment" in df.columns else 0
    _total_ads_perf = df["ads_perf"].abs().sum() if "ads_perf" in df.columns else 0
    ci1, ci2, ci3, ci4 = st.columns(4)
    ci1.metric("Выручка", r(total_rev))
    if total_partner or total_bonus:
        ci1.caption(f"из них: Программы {r(total_partner)}, Баллы {r(total_bonus)}")
    ci2.metric("Прибыль", r(total_prof))
    ci3.metric("Маржа", ru_pct(total_margin))
    ci4.metric("Расходы (комиссия + логистика)", r(total_comm + total_log))
    ci5, ci6, ci7, ci8 = st.columns(4)
    ci5.metric("Реклама (per-артикул)", r(_total_promo))
    ci6.metric("Эквайринг (per-артикул)", r(_total_acq))
    ci7.metric("Рассрочка (per-артикул)", r(_total_inst))
    ci8.metric("Налог (расчётный)", r(df["tax"].sum() if "tax" in df.columns else 0))
    if _total_ads_perf:
        st.metric("Реклама CPC/CPO по артикулам (Performance API)", r(_total_ads_perf))

    # Скачать Excel (форматированный — для просмотра)
    @st.cache_data
    def to_excel_display(df_in):
        """Форматированный Excel: числа с ₽ и пробелами. Артикулы — текст."""
        import io
        from openpyxl import load_workbook
        buf = io.BytesIO()
        df_in.to_excel(buf, index=False)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        # Находим колонку Артикул и принудительно ставим текстовый формат
        art_col_idx = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "Артикул":
                art_col_idx = col_idx
                break
        if art_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=art_col_idx, max_col=art_col_idx):
                for c in row:
                    c.number_format = "@"
                    if c.value is not None:
                        c.value = str(c.value)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @st.cache_data
    def to_excel_raw(df_in):
        """Сырой Excel: чистые числа без ₽, артикулы — текст. Для расчётов."""
        import io
        from openpyxl import load_workbook
        buf = io.BytesIO()
        # Копируем нужные колонки с чистыми числами
        raw_cols = {
            "article":     "Артикул",
            "name":        "Товар",
            "qty":         "Продано шт",
            "qty_ret":     "Возвращено шт",
            "revenue":     "Выручка",
            "partner_programs": "Программы партнёров",
            "bonus_points":     "Баллы за скидки",
            "cost_total":  "Себестоимость",
            "commission":  "Комиссия",
            "acquiring":   "Эквайринг",
            "tax":         "Налог",
            "logistics":   "Логистика",
            "promo":       "Реклама",
            "ads_cpc":     "Реклама CPC",
            "ads_cpo":     "Реклама CPO",
            "installment": "Рассрочка",
            "other_costs": "Прочие расходы",
            "profit":      "Прибыль",
            "margin_pct":  "Маржа %",
        }
        available = {k: v for k, v in raw_cols.items() if k in df_in.columns}
        raw_df = df_in[list(available.keys())].copy()
        raw_df = raw_df.rename(columns=available)
        # Артикул — строго текст
        raw_df["Артикул"] = raw_df["Артикул"].astype(str)
        raw_df.to_excel(buf, index=False)
        buf.seek(0)
        # Принудительно ставим текстовый формат на Артикул
        wb = load_workbook(buf)
        ws = wb.active
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "Артикул":
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for c in row:
                        c.number_format = "@"
                        if c.value is not None:
                            c.value = str(c.value)
                break
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        st.download_button(
            "⬇️ Скачать в Excel (отображение)",
            data=to_excel_display(display_df),
            file_name=f"ozon_pnl_{d_from}_{d_to}.xlsx",
            mime="application/vnd.ms-excel",
            help="С форматированием: числа с ₽, удобно для просмотра",
        )
    with btn_col2:
        st.download_button(
            "⬇️ Скачать данные (чистые числа)",
            data=to_excel_raw(df),
            file_name=f"ozon_pnl_raw_{d_from}_{d_to}.xlsx",
            mime="application/vnd.ms-excel",
            help="Чистые числа без ₽, артикулы как текст — для ВПР и расчётов в своём файле",
        )

    # ── Средняя логистика по артикулу ─────────────────────────────────────────
    st.divider()
    st.subheader("📦 Средняя логистика на единицу товара")
    if "logistics" in df.columns:
        log_df = df[df["logistics"].abs() > 0][["article", "name", "qty", "logistics"]].copy()
        log_df["logistics_abs"] = log_df["logistics"].abs()

        # Считаем уникальные отправления с ненулевой доставкой из raw_ops
        # Это точнее, чем qty_продаж — логистика может быть начислена раньше выручки
        raw_ops_log = st.session_state.get("raw_ops", [])
        sku_map_log  = st.session_state.get("sku_map_cache", {})
        shipment_sets: dict[str, set] = {}
        if raw_ops_log:
            for accrual in raw_ops_log:
                unit = accrual.get("unit_number", "")
                posting = accrual.get("posting") or {}
                for prod in (posting.get("products") or []):
                    if not isinstance(prod, dict):
                        continue
                    sku = str(prod.get("sku") or "")
                    offer_id = sku_map_log.get(sku, sku)
                    deliv = prod.get("delivery") or {}
                    total_v = float(((deliv.get("total_accrued") or {}).get("amount") or 0))
                    if total_v != 0 and unit:
                        shipment_sets.setdefault(offer_id, set()).add(unit)

        shipment_counts = {k: len(v) for k, v in shipment_sets.items()}
        if shipment_counts:
            log_df["отправлений"] = log_df["article"].map(shipment_counts).fillna(
                log_df["qty"].clip(lower=1)
            )
        else:
            log_df["отправлений"] = log_df["qty"].clip(lower=1)

        log_df["лог/отпр"] = (log_df["logistics_abs"] / log_df["отправлений"]).round(1)

        log_df = log_df.rename(columns={
            "article":      "Артикул",
            "name":         "Товар",
            "qty":          "Продано (финанс.)",
            "отправлений":  "Отправлений",
            "logistics_abs":"Логистика всего",
            "лог/отпр":     "Лог/отправление (ср.)",
        })
        log_df = log_df[[
            "Артикул", "Товар", "Продано (финанс.)", "Отправлений",
            "Логистика всего", "Лог/отправление (ср.)"
        ]].sort_values("Лог/отправление (ср.)", ascending=False)

        def color_logprice(val):
            if pd.isna(val): return ""
            if val >= 400: return "color: #F05B5B; font-weight: bold"
            if val >= 250: return "color: #F0A93D"
            return "color: #3DD68C"

        st.dataframe(
            log_df.style
            .format({
                "Продано (финанс.)":     lambda v: ru_float(v, 0),
                "Отправлений":           lambda v: ru_float(v, 0),
                "Логистика всего":       _fmt_rub,
                "Лог/отправление (ср.)": lambda v: ru_rub(v, 1),
            })
            .map(color_logprice, subset=["Лог/отправление (ср.)"]),
            use_container_width=True, hide_index=True,
        )
        st.caption(
            "🔴 ≥400 ₽ — высокая логистика (проверь упаковку/склад) · 🟡 250–400 ₽ · 🟢 <250 ₽  \n"
            "**Отправлений** — уникальные заказы с начисленной логистикой. "
            "**Продано (финанс.)** — продажи с зачисленной выручкой (могут не совпадать из-за задержки начислений)."
        )

def page_stocks():
    # ── Остатки и оборачиваемость ──────────────────────────────────────────────
    st.subheader("📦 Остатки и оборачиваемость")
    stocks_data = st.session_state.get("stocks_data", [])

    if not stocks_data:
        st.info("Остатки загружаются вместе с данными. Нажми **Загрузить данные** в боковом меню.")
    else:
        # Данные из /v1/analytics/turnover/stocks
        # Поля: offer_id, name, sku, current_stock, ads, idc, idc_grade, turnover, turnover_grade
        stocks_df = pd.DataFrame(stocks_data)

        # Нужные колонки — добавляем если отсутствуют
        for col, default in [("offer_id",""), ("name",""), ("current_stock",0),
                              ("ads",0.0), ("idc",0.0), ("idc_grade",""), ("turnover",0.0), ("turnover_grade","")]:
            if col not in stocks_df.columns:
                stocks_df[col] = default

        stocks_df["offer_id"] = stocks_df["offer_id"].astype(str)

        # Переводим оценки в читаемый вид
        stocks_df["Статус запаса"] = stocks_df["idc_grade"].map(TURNOVER_GRADE_RU).fillna(stocks_df["idc_grade"])
        stocks_df["Статус оборач."] = stocks_df["turnover_grade"].map(TURNOVER_GRADE_RU).fillna(stocks_df["turnover_grade"])

        # Продажи из P&L за период (для сравнения)
        loaded_period = st.session_state.get("loaded_period")
        if loaded_period:
            p_from, p_to = loaded_period
            period_days_calc = max(1, (p_to - p_from).days + 1)
        else:
            period_days_calc = max(1, days)

        sales_df = df[df["qty"] > 0][["article", "qty"]].copy()
        sales_df = sales_df.rename(columns={"article": "offer_id"})
        stocks_df = stocks_df.merge(sales_df, on="offer_id", how="left")
        stocks_df["qty"] = stocks_df["qty"].fillna(0)
        stocks_df["Продано (P&L)"] = stocks_df["qty"].astype(int)

        # Сортировка — сначала критичные
        grade_order = {
            "GRADES_CRITICAL": 0, "GRADES_RED": 1, "GRADES_YELLOW": 2,
            "GRADES_GREEN": 3, "GRADES_NOSALES": 4, "GRADES_NONE": 5, "": 6,
        }
        stocks_df["_sort"] = stocks_df["idc_grade"].map(grade_order).fillna(6)
        stocks_df = stocks_df.sort_values(["_sort", "idc"])

        # Итоговая таблица
        out_df = stocks_df[[
            "offer_id", "name", "current_stock", "ads", "idc", "turnover",
            "Статус запаса", "Статус оборач.", "Продано (P&L)"
        ]].rename(columns={
            "offer_id":      "Артикул",
            "name":          "Товар",
            "current_stock": "Остаток",
            "ads":           "Ср. прод/день (Ozon)",
            "idc":           "Дней запаса",
            "turnover":      "Оборач. (дни)",
        })

        def color_idc(val):
            if pd.isna(val): return ""
            if val == 0: return "color: #F05B5B; font-weight: bold"
            if val < 28: return "color: #F05B5B; font-weight: bold"
            if val < 56: return "color: #F0A93D"
            return "color: #3DD68C"

        st.dataframe(
            out_df.style
            .format({
                "Остаток":               lambda v: ru_float(v, 0),
                "Ср. прод/день (Ozon)":  lambda v: ru_float(v, 2),
                "Дней запаса":           lambda v: ru_float(v, 1) if not pd.isna(v) else "—",
                "Оборач. (дни)":         lambda v: ru_float(v, 1) if not pd.isna(v) else "—",
                "Продано (P&L)":         lambda v: ru_float(v, 0),
            })
            .map(color_idc, subset=["Дней запаса"]),
            use_container_width=True,
            height=500,
            hide_index=True,
        )

        st.caption(
            "**Дней запаса** и **Ср. прод/день** — данные Ozon за 60 дней. "
            "**Оборач.** — фактическая оборачиваемость в днях. "
            "**Продано (P&L)** — за выбранный период из P&L."
        )

        # KPI-сводка
        grade_counts = stocks_df["idc_grade"].value_counts()
        ks1, ks2, ks3, ks4, ks5 = st.columns(5)
        ks1.metric("🔴 Критично",   int(grade_counts.get("GRADES_CRITICAL", 0)) + int(grade_counts.get("GRADES_RED", 0)))
        ks2.metric("🟡 Средне",     int(grade_counts.get("GRADES_YELLOW", 0)))
        ks3.metric("🟢 Хорошо",     int(grade_counts.get("GRADES_GREEN", 0)))
        ks4.metric("⚪ Нет продаж", int(grade_counts.get("GRADES_NOSALES", 0)))
        ks5.metric("⏳ Нет остатка", int(grade_counts.get("GRADES_NONE", 0)))

        # Предупреждения по критичным позициям
        critical_items = out_df[out_df["Дней запаса"] < 28].copy()
        critical_items = critical_items[critical_items["Дней запаса"] > 0]
        if not critical_items.empty:
            st.warning(
                f"⚠️ {len(critical_items)} товаров с запасом менее 28 дней. "
                "Требуется пополнение или включать рекламу только после завоза."
            )
            with st.expander("Критичные позиции", expanded=True):
                st.dataframe(
                    critical_items[["Артикул", "Товар", "Остаток", "Ср. прод/день (Ozon)", "Дней запаса", "Статус запаса"]],
                    use_container_width=True, hide_index=True,
                )

        # Скачать
        @st.cache_data
        def stocks_to_excel(df_in):
            import io
            buf = io.BytesIO()
            df_in.to_excel(buf, index=False)
            return buf.getvalue()

        st.download_button(
            "⬇️ Скачать остатки в Excel",
            data=stocks_to_excel(out_df),
            file_name="ozon_stocks.xlsx",
            mime="application/vnd.ms-excel",
        )

def page_details():
    st.subheader("🔍 Детализация по заказам")

    raw_ops = st.session_state.get("raw_ops", [])
    sku_map_cache = st.session_state.get("sku_map_cache", {})

    if not raw_ops:
        st.info("Данные детализации доступны после загрузки в режиме 'Текущий месяц' или 'Произвольные даты'.")
        st.stop()

    # ── Диагностика структуры API ──────────────────────────────────────────────
    with st.expander("🔧 Диагностика API", expanded=False):
        import json

        # 1. Категории начислений и их количество
        cat_counts: dict[str, int] = {}
        cat_has_item_fees: dict[str, int] = {}
        for _a in raw_ops:
            _cat = _a.get("accrued_category", "UNKNOWN")
            cat_counts[_cat] = cat_counts.get(_cat, 0) + 1
            _ifs = _a.get("item_fees")
            if _ifs and isinstance(_ifs, dict) and (_ifs.get("fees") or []):
                cat_has_item_fees[_cat] = cat_has_item_fees.get(_cat, 0) + 1

        st.subheader("Категории начислений")
        _cat_rows = [{"Категория": k, "Кол-во": v,
                      "С item_fees": cat_has_item_fees.get(k, 0)} for k, v in sorted(cat_counts.items())]
        st.dataframe(pd.DataFrame(_cat_rows), hide_index=True, use_container_width=True)

        # 2. Логистика: сопоставленная vs несопоставленная (numeric SKU без offer_id)
        st.subheader("Логистика: сопоставлено / несопоставлено")
        _logi_matched = 0.0
        _logi_unmatched = 0.0
        _unmatched_skus: dict[str, float] = {}
        for _a in raw_ops:
            _p = _a.get("posting") or {}
            for _pr in (_p.get("products") or []):
                if not isinstance(_pr, dict): continue
                _sku = str(_pr.get("sku") or "")
                _v = float((((_pr.get("delivery") or {}).get("total_accrued") or {}).get("amount") or 0))
                if not _v: continue
                if sku_map_cache.get(_sku):
                    _logi_matched += _v
                else:
                    _logi_unmatched += _v
                    _unmatched_skus[_sku] = _unmatched_skus.get(_sku, 0.0) + _v
        st.metric("Сопоставлено (есть offer_id)", f"{abs(_logi_matched):,.0f} ₽")
        st.metric("Несопоставлено (нет в справочнике)", f"{abs(_logi_unmatched):,.0f} ₽")
        if _unmatched_skus:
            st.caption("Несопоставленные SKU:")
            st.dataframe(pd.DataFrame([{"SKU": k, "Логистика": v}
                                        for k, v in _unmatched_skus.items()]),
                         hide_index=True, use_container_width=True)

        # 3. item_fees: тип начисления и суммы
        st.subheader("Суммы по item_fees (type_id / accrual_id)")
        _fee_totals: dict[int, float] = {}
        for _a in raw_ops:
            _ifs = _a.get("item_fees") or {}
            for _sf in (_ifs.get("fees") or []):
                if not isinstance(_sf, dict): continue
                for _f in (_sf.get("fees") or []):
                    if not isinstance(_f, dict): continue
                    _tid = _get_type_id(_f)
                    _amt = float(((_f.get("accrued") or {}).get("amount") or 0))
                    if _amt: _fee_totals[_tid] = _fee_totals.get(_tid, 0.0) + _amt
        if _fee_totals:
            st.dataframe(pd.DataFrame([
                {"type_id": k, "Название": TYPE_NAMES.get(k, "⚠️ Неизвестно"), "Сумма": v}
                for k, v in sorted(_fee_totals.items(), key=lambda x: x[1])
            ]), hide_index=True, use_container_width=True)
        else:
            st.info("item_fees пусты для всех начислений в этом периоде")

        # 4. Сырой пример POSTING-начисления
        st.subheader("Пример POSTING (products)")
        _sample = next(
            (a for a in raw_ops if a.get("posting") and (a["posting"].get("products") or [])),
            raw_ops[0] if raw_ops else {}
        )
        st.json(_sample, expanded=2)

        # 4b. Поиск конкретного начисления по ID — чтобы найти поле "Баллы за скидки" /
        # "Программы партнёров" (эти суммы есть в выгрузке XLSX, но НЕ парсятся сейчас в коде —
        # в transactions_to_df() читается только commission.sale_amount / sale_commission / seller_price).
        st.subheader("Найти начисление по ID (для поиска поля 'Баллы за скидки' / 'Программы партнёров')")
        st.caption(
            "По умолчанию — заказ 44329526-0299-1: в выгрузке 'Отчёт по начислениям.xlsx' у него "
            "Выручка=1198.10, Программы партнёров=11.98, Баллы за скидки=1248.92. "
            "Разверни posting → products → [0] → commission и найди ключи с такими суммами."
        )
        _search_id = st.text_input("ID начисления / unit_number", value="44329526-0299-1")
        if _search_id:
            _matches = [a for a in raw_ops if str(a.get("unit_number", "")) == _search_id.strip()]
            if _matches:
                st.write(f"Найдено начислений: {len(_matches)}")
                for _m in _matches:
                    st.json(_m, expanded=4)
            else:
                st.warning("Начисление с таким ID не найдено в загруженном периоде.")

        # 5. Итоговые суммы в финальном df (после enrich_with_cost)
        st.subheader("Итоги в финальном df (после enrich_with_cost)")
        _fdf = st.session_state.get("df")
        if _fdf is not None and not _fdf.empty:
            _d5c1, _d5c2, _d5c3, _d5c4 = st.columns(4)
            _d5c1.metric("Реклама (promo) — сумма", f"{_fdf['promo'].abs().sum():,.0f} ₽".replace(",", " ") if "promo" in _fdf.columns else "нет колонки")
            _d5c2.metric("Эквайринг (acquiring) — сумма", f"{_fdf['acquiring'].abs().sum():,.0f} ₽".replace(",", " ") if "acquiring" in _fdf.columns else "нет колонки")
            _d5c3.metric("Рассрочка (installment) — сумма", f"{_fdf['installment'].abs().sum():,.0f} ₽".replace(",", " ") if "installment" in _fdf.columns else "нет колонки")
            _d5c4.metric("Логистика (logistics) — сумма", f"{_fdf['logistics'].abs().sum():,.0f} ₽".replace(",", " ") if "logistics" in _fdf.columns else "нет колонки")
            st.caption("Если Реклама и Эквайринг = 0 — item_fees не дошли до таблицы (нужно проверить сопоставление SKU). Если > 0 — данные есть и отображаются в колонках.")

            st.markdown("**Performance API (реклама CPC/CPO по артикулам):**")
            _perf_meta_diag = st.session_state.get("perf_ad_meta", {}) or {}
            _d5p1, _d5p2, _d5p3, _d5p4 = st.columns(4)
            _d5p1.metric("Реклама CPC/CPO (ads_perf) — сумма", f"{_fdf['ads_perf'].abs().sum():,.0f} ₽".replace(",", " ") if "ads_perf" in _fdf.columns else "нет колонки")
            _d5p2.metric("Найдено SKU с расходом", str(len(_perf_meta_diag.get("by_sku", {}))))
            _d5p3.metric("CPC: сумма / кампаний", f"{_perf_meta_diag.get('cpc_total', 0):,.0f} ₽ / {_perf_meta_diag.get('cpc_campaigns_used', 0)}".replace(",", " "))
            _d5p4.metric("CPO: сумма / источник", f"{_perf_meta_diag.get('cpo_total', 0):,.0f} ₽ / {_perf_meta_diag.get('cpo_source') or '—'}".replace(",", " "))
            if _perf_meta_diag.get("error"):
                st.error(f"Ошибка Performance API при последней загрузке: {_perf_meta_diag['error']}")
            elif not _perf_meta_diag.get("fetched"):
                st.info("Performance API не запрашивался — не заполнены Performance Client-ID/Secret в боковом меню.")
            else:
                st.caption(
                    f"CPC: {'ок' if _perf_meta_diag.get('cpc_ok') else '⚠️ не удалось'} "
                    f"({_perf_meta_diag.get('cpc_campaigns_used', 0)} кампаний использовано, "
                    f"{len(_perf_meta_diag.get('cpc_campaigns_failed', []))} с ошибкой). "
                    f"CPO: {'ок' if _perf_meta_diag.get('cpo_ok') else '⚠️ не удалось'} "
                    f"(источник: {_perf_meta_diag.get('cpo_source') or '—'})."
                    + (f" Ошибка CPO: {_perf_meta_diag['cpo_error']}" if _perf_meta_diag.get('cpo_error') else "")
                )
            if _perf_meta_diag.get("raw_campaigns_sample"):
                st.caption(
                    f"Сырые кампании из GET /api/client/campaign (всего "
                    f"{_perf_meta_diag.get('raw_campaigns_count', 0)}, показаны первые "
                    f"{len(_perf_meta_diag['raw_campaigns_sample'])}) — реальные значения "
                    f"paymentType/advObjectType, по которым идёт фильтр CPC/CPO:"
                )
                st.dataframe(
                    pd.DataFrame(_perf_meta_diag["raw_campaigns_sample"]),
                    use_container_width=True, hide_index=True,
                )
            if _perf_meta_diag.get("cpo_raw_sample"):
                st.caption("Сырой пример строк CPO-отчёта (для диагностики схемы ответа API):")
                st.json(_perf_meta_diag["cpo_raw_sample"])
            if _perf_meta_diag.get("cpc_error_samples"):
                st.caption(
                    "Реальные ошибки API по проблемным CPC-кампаниям (до 5 примеров) — "
                    "статус и текст ответа Ozon, а не догадки:"
                )
                st.dataframe(
                    pd.DataFrame(_perf_meta_diag["cpc_error_samples"]),
                    use_container_width=True, hide_index=True,
                )
        else:
            st.warning("df ещё не загружен — сначала нажмите «Загрузить данные»")

        # 6. Себестоимость: диагностика cost_map ↔ df["article"]
        st.subheader("Себестоимость: диагностика")
        _cm_debug = st.session_state.get("_cost_map_debug", {})
        _fdf_cost = st.session_state.get("df")
        _dc1, _dc2, _dc3 = st.columns(3)
        _dc1.metric("Записей в cost_map (XLSX)", str(len(_cm_debug)))
        _dc2.metric("Записей в sku_map_cache", str(len(sku_map_cache)))
        if _fdf_cost is not None and not _fdf_cost.empty:
            _articles_list = _fdf_cost["article"].tolist()
            _matches_count = sum(1 for a in _articles_list if a in _cm_debug)
            _dc3.metric("Совпадений article↔cost_map", str(_matches_count))
            if _cm_debug:
                st.caption("Первые 5 ключей cost_map (из XLSX):")
                st.code(str(list(_cm_debug.keys())[:5]))
            st.caption("Первые 5 значений df['article'] (из API):")
            st.code(str(_articles_list[:5]))
            if _matches_count == 0 and _cm_debug:
                st.error("❌ Нет совпадений: артикулы в XLSX не совпадают с df['article']. "
                         "Если df['article'] — числа, значит sku_map пуст и SKU не были сопоставлены с offer_id.")
        else:
            _dc3.metric("Совпадений article↔cost_map", "—")

        # 7. Справочник типов начислений: сырой ответ
        st.subheader("Справочник типов начислений (raw API)")
        if client_id and api_key:
            with st.spinner("Запрос /v1/finance/accrual/types..."):
                _raw_accrual_types = api_post("/v1/finance/accrual/types", {}, client_id, api_key)
            if _raw_accrual_types:
                st.json(_raw_accrual_types)
            else:
                st.warning("Пустой ответ от /v1/finance/accrual/types")
        else:
            st.info("Введи API-ключи чтобы увидеть справочник типов")

        # 8. Диагностика FBO/FBS list — почему sku_map пустой?
        st.subheader("Диагностика: FBO/FBS list (почему sku_map пустой?)")
        if client_id and api_key:
            import datetime as _dt
            _today = _dt.date.today()
            _since = (_today - _dt.timedelta(days=30)).strftime("%Y-%m-%d")
            _to_s  = _today.strftime("%Y-%m-%d")
            for _schema, _path in [("FBO /v3", "/v3/posting/fbo/list"), ("FBS /v4", "/v4/posting/fbs/list")]:
                with st.spinner(f"Запрос {_schema}..."):
                    _body = {
                        "dir": "ASC",
                        "filter": {"since": _since + "T00:00:00.000Z", "to": _to_s + "T23:59:59.000Z"},
                        "limit": 5, "offset": 0,
                        "with": {"financial_data": False, "analytics_data": False}
                    }
                    _resp = api_post(_path, _body, client_id, api_key)
                st.markdown(f"**{_schema}** (`{_since}` → `{_to_s}`):")
                if _resp is None:
                    st.error(f"{_schema}: ответ None — скорее всего ошибка авторизации или API")
                else:
                    _result = _resp.get("result", {})
                    _postings = _result if isinstance(_result, list) else _result.get("postings", [])
                    st.write(f"Найдено заказов (первые 5): {len(_postings)}")
                    if _postings:
                        _sample = _postings[0]
                        _prods = (_sample.get("products") or [])[:2]
                        st.write("Пример товаров:", _prods)
                    else:
                        st.json(_resp)  # покажем полный ответ если пусто
        else:
            st.info("Введи API-ключи")

    # Обратный справочник: offer_id → список numeric SKU
    reverse_map: dict[str, list[str]] = {}
    for num_sku, offer_id in sku_map_cache.items():
        reverse_map.setdefault(offer_id, []).append(str(num_sku))

    articles = sorted(df["article"].unique().tolist())
    sel_article = st.selectbox("Артикул", articles)
    sel_skus = set(reverse_map.get(sel_article, [sel_article]))

    detail_rows = []
    typeid_totals: dict[int, float] = {}

    for accrual in raw_ops:
        date_str = accrual.get("_date", "")
        unit = accrual.get("unit_number", "")
        category = accrual.get("accrued_category", "")

        # Posting — продажи
        posting = accrual.get("posting") or {}
        for prod in (posting.get("products") or []):
            if not isinstance(prod, dict):
                continue
            if str(prod.get("sku") or "") not in sel_skus:
                continue

            comm = prod.get("commission") or {}
            deliv = prod.get("delivery") or {}
            services = deliv.get("services") or []

            # sale_amount/seller_price = ПОЛНАЯ сумма (Выручка+Программы+Баллы) — см. диагностику от 02.08.2026.
            # Для деталки показываем настоящую Выручку (sale_price) + отдельно Программы/Баллы.
            revenue_v = float(((comm.get("sale_price") or {}).get("amount") or 0))
            partner_v = float(((comm.get("coinvestment") or {}).get("amount") or 0))
            bonus_v   = float(((comm.get("bonus") or {}).get("amount") or 0))
            commission_v = float(((comm.get("sale_commission") or {}).get("amount") or 0))
            total_deliv_v = float(((deliv.get("total_accrued") or {}).get("amount") or 0))

            for svc in services:
                if not isinstance(svc, dict):
                    continue
                tid = _get_type_id(svc)
                amt = float(((svc.get("accrued") or {}).get("amount") or 0))
                src = "delivery.services"
                typeid_totals[(tid, src)] = typeid_totals.get((tid, src), 0) + amt
                detail_rows.append({
                    "Дата": date_str,
                    "Заказ": unit,
                    "Схема": posting.get("delivery_schema", ""),
                    "fee_type_id": tid,
                    "Название": TYPE_NAMES.get(tid, f"delivery type_id={tid}"),
                    "Сумма": amt,
                    "Выручка": revenue_v,
                    "Программы партнёров": partner_v,
                    "Баллы за скидки": bonus_v,
                    "Комиссия": commission_v,
                    "Доставка (итого)": total_deliv_v,
                })
                revenue_v = 0  # не дублируем выручку на каждый сервис одного заказа
                partner_v = 0
                bonus_v = 0
                commission_v = 0
                total_deliv_v = 0

            if not services:
                detail_rows.append({
                    "Дата": date_str,
                    "Заказ": unit,
                    "Схема": posting.get("delivery_schema", ""),
                    "fee_type_id": None,
                    "Название": "—",
                    "Сумма": 0,
                    "Выручка": revenue_v,
                    "Программы партнёров": partner_v,
                    "Баллы за скидки": bonus_v,
                    "Комиссия": commission_v,
                    "Доставка (итого)": total_deliv_v,
                })

        # Item fees — прочие начисления по SKU (эквайринг, реклама и т.п.)
        item_fees_block = accrual.get("item_fees") or {}
        for sku_fees in (item_fees_block.get("fees") or []):
            if not isinstance(sku_fees, dict):
                continue
            if str(sku_fees.get("sku") or "") not in sel_skus:
                continue
            for fee in (sku_fees.get("fees") or []):
                if not isinstance(fee, dict):
                    continue
                tid = _get_type_id(fee)
                amt = float(((fee.get("accrued") or {}).get("amount") or 0))
                src = "item_fees"
                typeid_totals[(tid, src)] = typeid_totals.get((tid, src), 0) + amt
                detail_rows.append({
                    "Дата": date_str,
                    "Заказ": unit,
                    "Схема": TYPE_NAMES.get(tid, f"ITEM type_id={tid}"),
                    "fee_type_id": tid,
                    "Сумма": amt,
                    "Выручка": 0,
                    "Комиссия": 0,
                    "Доставка (итого)": 0,
                })

    if detail_rows:
        det_df = pd.DataFrame(detail_rows).fillna(0)
        num_cols = [c for c in det_df.columns if c not in ("Дата", "Заказ", "Схема", "fee_type_id")]
        for c in num_cols:
            det_df[c] = pd.to_numeric(det_df[c], errors="coerce").fillna(0)

        # Убираем нулевые числовые строки из вывода
        mask = det_df[num_cols].abs().sum(axis=1) > 0
        det_shown = det_df[mask]
        st.markdown(f"**Строк с данными: {mask.sum()}**")

        # Подсчёт компонентов логистики для объяснения
        if "Доставка (итого)" in det_shown.columns and "Заказ" in det_shown.columns:
            total_deliv_sum = det_shown["Доставка (итого)"].sum()
            # Считаем уникальные заказы, у которых была ненулевая логистика
            orders_with_logi = det_shown[det_shown["Доставка (итого)"].abs() > 0]["Заказ"].nunique()
            if orders_with_logi > 0 and total_deliv_sum != 0:
                avg_logi = abs(total_deliv_sum) / orders_with_logi
                st.caption(
                    f"Логистика всего за период: **{ru_rub(abs(total_deliv_sum), 2)}** · "
                    f"Отправлений с логистикой: **{orders_with_logi} шт** · "
                    f"Ср. на отправление: **{ru_rub(avg_logi, 1)}** "
                    f"= Доставка (итого) / кол-во отправлений"
                )

        st.dataframe(det_shown.style.format(
            {c: lambda v: ru_float(v, 2) for c in num_cols}
        ), use_container_width=True, height=400)

        # Сводка начислений по артикулу
        st.subheader("Начисления по выбранному артикулу")
        if typeid_totals:
            known_item_ids = {ACQUIRING_TYPE_ID, INSTALLMENT_TYPE_ID} | PROMO_TYPE_IDS
            tid_rows = []
            for (tid, src), amt in sorted(typeid_totals.items(), key=lambda x: (x[0][0] or 0)):
                on_dash = tid in known_item_ids or src == "delivery.services"
                tid_rows.append({
                    "Код (type_id)": tid,
                    "Название": TYPE_NAMES.get(tid, "⚠️ Неизвестно"),
                    "Источник": src,
                    "Сумма": amt,
                    "Учтён": "✅" if on_dash else "❌",
                })
            tid_df = pd.DataFrame(tid_rows)
            st.dataframe(
                tid_df.style.format({"Сумма": lambda v: ru_float(v, 2)}),
                use_container_width=True,
                hide_index=True,
            )
    else:
        st.warning(f"Нет данных для артикула {sel_article} в загруженном периоде.")

    # ── Все начисления кабинета за период ─────────────────────────────────────
    st.divider()
    st.subheader("Все начисления кабинета за период")

    cab_totals: dict[tuple, float] = {}
    for accrual in raw_ops:
        # delivery.services (по всем заказам)
        posting = accrual.get("posting") or {}
        for prod in (posting.get("products") or []):
            if not isinstance(prod, dict):
                continue
            for svc in ((prod.get("delivery") or {}).get("services") or []):
                if not isinstance(svc, dict):
                    continue
                tid = _get_type_id(svc)
                amt = float(((svc.get("accrued") or {}).get("amount") or 0))
                cab_totals[(tid, "delivery.services")] = cab_totals.get((tid, "delivery.services"), 0) + amt

        # item_fees (по всем SKU)
        for sku_fees in ((accrual.get("item_fees") or {}).get("fees") or []):
            if not isinstance(sku_fees, dict):
                continue
            for fee in (sku_fees.get("fees") or []):
                if not isinstance(fee, dict):
                    continue
                tid = _get_type_id(fee)
                amt = float(((fee.get("accrued") or {}).get("amount") or 0))
                cab_totals[(tid, "item_fees")] = cab_totals.get((tid, "item_fees"), 0) + amt

        # non_item_fee — общие сборы на уровне заказа (хранение и т.п.)
        nif = accrual.get("non_item_fee")
        if isinstance(nif, dict) and _get_type_id(nif) is not None:
            tid = _get_type_id(nif)
            amt = float(((nif.get("accrued") or {}).get("amount") or 0))
            cab_totals[(tid, "non_item_fee")] = cab_totals.get((tid, "non_item_fee"), 0) + amt

    if cab_totals:
        known_ids = {ACQUIRING_TYPE_ID, INSTALLMENT_TYPE_ID, 29, 32} | PROMO_TYPE_IDS | set(STORE_COST_GROUPS.keys())
        cab_rows = []
        for (tid, src), amt in sorted(cab_totals.items(), key=lambda x: x[1]):
            if amt == 0:
                continue
            name = TYPE_NAMES.get(tid, "⚠️ Неизвестно")
            on_dashboard = tid in known_ids or src == "delivery.services"
            cab_rows.append({
                "Код (type_id)": tid,
                "Название": name,
                "Источник": src,
                "Сумма": amt,
                "Учтён": "✅" if on_dashboard else "❌",
            })
        cab_df = pd.DataFrame(cab_rows)

        # Сначала показываем нераспознанные — они самые важные
        unrecognized = cab_df[cab_df["Учтён"] == "❌"].copy()
        recognized   = cab_df[cab_df["Учтён"] == "✅"].copy()

        unrecognized_expenses = unrecognized[unrecognized["Сумма"] < 0]
        unrecognized_income   = unrecognized[unrecognized["Сумма"] > 0]
        total_unrecognized_exp = unrecognized_expenses["Сумма"].sum()
        total_unrecognized_inc = unrecognized_income["Сумма"].sum()

        if not unrecognized.empty:
            st.error(
                f"⚠️ Нераспознанные начисления: расходы **{total_unrecognized_exp:,.0f} ₽**, "
                f"доходы **+{total_unrecognized_inc:,.0f} ₽** — не вошли в P&L."
            )
            with st.expander("❌ Нераспознанные начисления (не в P&L) — развернуть", expanded=True):
                st.dataframe(
                    unrecognized.style.format({"Сумма": lambda v: ru_float(v, 2)}),
                    use_container_width=True, hide_index=True,
                )
                st.caption(
                    "Эти начисления есть в API, но дашборд их не классифицировал. "
                    "Сообщи, в какую категорию отнести каждый код — добавим в формулу."
                )
        else:
            st.success("✅ Все начисления распознаны и учтены в P&L.")

        with st.expander("✅ Учтённые начисления (полный список)", expanded=False):
            st.dataframe(
                recognized.style.format({"Сумма": lambda v: ru_float(v, 2)}),
                use_container_width=True, hide_index=True,
                height=400,
            )

def page_pnl():
    """
    Страница «P&L» (Task #20) — сводка по периодам (день/неделя/месяц) с
    раскрывающимися группами столбцов (streamlit-aggrid), по образцу скриншотов
    стороннего сервиса аналитики, которые прислал пользователь 03.08.2026.

    Отдельная страница, НЕ замена «По артикулам» — по явному решению пользователя
    (there were 3 варианта, выбран «Отдельная страница»). Формулы Операционной/
    Чистой прибыли — интерпретация макета, см. docstring build_period_pnl().

    История периодов пока НЕ сохраняется между загрузками (Task #17/#19, Google
    Sheets — заблокировано, ждём service account JSON от пользователя). Каждая
    загрузка показывает P&L только за выбранный в сайдбаре период.
    """
    st.subheader("📈 P&L по периодам")

    raw_ops_pnl = st.session_state.get("raw_ops", [])
    if not raw_ops_pnl:
        st.info("👈 Сначала загрузи данные в боковом меню («Загрузить данные»).")
        return

    cost_map_pnl = st.session_state.get("_cost_map_debug", {})
    perf_meta_pnl = st.session_state.get("perf_ad_meta", {}) or {}
    _use_perf_pnl = bool(use_perf_ads)
    cpc_by_date_pnl = perf_meta_pnl.get("cpc_by_date", {}) if _use_perf_pnl else {}
    cpo_by_date_pnl = perf_meta_pnl.get("cpo_by_date", {}) if _use_perf_pnl else {}

    daily_pnl = transactions_to_daily_pnl(raw_ops_pnl, cost_map_pnl, cpc_by_date_pnl, cpo_by_date_pnl)
    if daily_pnl.empty:
        st.info("Нет данных о начислениях за выбранный период.")
        return
    store_daily_pnl = collect_store_costs_daily(raw_ops_pnl)

    gran_map = {"По дням": "D", "По неделям": "W", "По месяцам": "M"}
    gran_label = st.radio("Группировка периода", list(gran_map.keys()), index=1, horizontal=True)
    freq = gran_map[gran_label]

    period_df = build_period_pnl(
        daily_pnl, store_daily_pnl, freq,
        perf_replaces_cpc=_perf_replaces_cpc, perf_replaces_cpo=_perf_replaces_cpo,
    )
    if period_df.empty:
        st.info("Нет данных за выбранный период.")
        return

    display = pd.DataFrame({
        "period":             period_df["period_label"],
        "qty_total":          period_df["qty_total"],
        "qty":                period_df["qty"],
        "qty_ret":            period_df["qty_ret"],
        "revenue":            period_df["revenue"],
        "total_income":       period_df["total_income"],
        "commission_abs":     period_df["commission_abs"],
        "commission_pct":     period_df["commission_pct"],
        "cost_total":         period_df["cost_total"],
        "logistics_abs":      period_df["logistics_abs"],
        "acquiring":          period_df["acquiring"],
        "installment":        period_df["installment"],
        "other_costs_abs":    period_df["other_costs_abs"],
        "other_store_abs":    period_df["other_store_abs"],
        "mp_expenses_total":  period_df["mp_expenses_total"],
        "marketing_total":    period_df["marketing_total"],
        "tax":                period_df["tax"],
        "operating_profit":   period_df["operating_profit"],
        "roi_operating_pct":  period_df["roi_operating_pct"],
        "net_profit":         period_df["net_profit"],
        "roi_net_pct":        period_df["roi_net_pct"],
    })

    if _AGGRID_AVAILABLE:
        _rub_fmt = JsCode("""
            function(params) {
                if (params.value === null || params.value === undefined || isNaN(params.value)) return '—';
                var v = Math.round(params.value);
                var s = Math.abs(v).toLocaleString('ru-RU').replace(/,/g, ' ');
                return (v < 0 ? '-' : '') + s + ' ₽';
            }
        """)
        _pct_fmt = JsCode("""
            function(params) {
                if (params.value === null || params.value === undefined || isNaN(params.value)) return '—';
                return params.value.toFixed(1).replace('.', ',') + ' %';
            }
        """)
        _int_fmt = JsCode("""
            function(params) {
                if (params.value === null || params.value === undefined || isNaN(params.value)) return '—';
                return Math.round(params.value).toLocaleString('ru-RU').replace(/,/g, ' ');
            }
        """)
        _profit_style = JsCode("""
            function(params) {
                if (params.value === null || params.value === undefined) return {};
                return { color: params.value >= 0 ? '#3DD68C' : '#F05B5B', fontWeight: 'bold' };
            }
        """)

        def _leaf(field, header, fmt=_rub_fmt, width=110, style=None):
            d = {"field": field, "headerName": header, "width": width, "valueFormatter": fmt}
            if style:
                d["cellStyle"] = style
            return d

        column_defs = [
            {"field": "period", "headerName": "Период", "pinned": "left", "width": 150},
            {"headerName": "Продажи", "children": [
                _leaf("qty_total", "Всего", _int_fmt, 90),
                _leaf("qty", "Продажи", _int_fmt, 90),
                _leaf("qty_ret", "Возвраты", _int_fmt, 90),
            ]},
            _leaf("revenue", "Выручка", width=120),
            _leaf("total_income", "Доход с продаж покупателям", width=160),
            {"headerName": "Комиссия", "children": [
                _leaf("commission_abs", "Сумма", width=110),
                _leaf("commission_pct", "Процент", _pct_fmt, 90),
            ]},
            _leaf("cost_total", "Себестоимость", width=130),
            {"headerName": "Расходы внутри МП", "children": [
                _leaf("logistics_abs", "Логистика", width=110),
                _leaf("acquiring", "Эквайринг", width=110),
                _leaf("installment", "Рассрочка", width=110),
                _leaf("other_costs_abs", "Прочее (по SKU)", width=130),
                _leaf("other_store_abs", "Прочее (магазин)", width=140),
                _leaf("mp_expenses_total", "Итого", width=120),
            ]},
            _leaf("marketing_total", "Маркетинг", width=120),
            _leaf("tax", "Налог", width=100),
            {"headerName": "Операционная прибыль", "children": [
                _leaf("operating_profit", "Сумма", width=130, style=_profit_style),
                _leaf("roi_operating_pct", "ROI", _pct_fmt, 90),
            ]},
            {"headerName": "Чистая прибыль", "children": [
                _leaf("net_profit", "Сумма", width=130, style=_profit_style),
                _leaf("roi_net_pct", "ROI", _pct_fmt, 90),
            ]},
        ]

        grid_options = {
            "columnDefs": column_defs,
            "defaultColDef": {"resizable": True, "sortable": False, "suppressMovable": True},
            "suppressColumnVirtualisation": True,
        }

        AgGrid(
            display,
            gridOptions=grid_options,
            allow_unsafe_jscode=True,
            theme="streamlit",
            height=min(80 + 42 * len(display), 600),
            show_toolbar=False,
            show_search=False,
            show_download_button=False,
        )
    else:
        st.warning(
            "⚠️ Пакет streamlit-aggrid ещё не подхватился на сервере — показываю обычную таблицу "
            "без раскрывающихся групп столбцов. Обычно помогает «Reboot app» в Streamlit Cloud "
            "после первого деплоя с обновлённым requirements.txt."
        )
        rename = {
            "period": "Период", "qty_total": "Всего шт", "qty": "Продано шт", "qty_ret": "Возвращено шт",
            "revenue": "Выручка", "total_income": "Доход с продаж покупателям",
            "commission_abs": "Комиссия, ₽", "commission_pct": "Комиссия, %",
            "cost_total": "Себестоимость",
            "logistics_abs": "Логистика", "acquiring": "Эквайринг", "installment": "Рассрочка",
            "other_costs_abs": "Прочее (SKU)", "other_store_abs": "Прочее (магазин)",
            "mp_expenses_total": "Расходы внутри МП, итого",
            "marketing_total": "Маркетинг", "tax": "Налог",
            "operating_profit": "Операционная прибыль", "roi_operating_pct": "ROI операц., %",
            "net_profit": "Чистая прибыль", "roi_net_pct": "ROI чистая, %",
        }
        fallback = display.rename(columns=rename)
        skip_rub = {"Период", "Всего шт", "Продано шт", "Возвращено шт", "Комиссия, %", "ROI операц., %", "ROI чистая, %"}
        fmt_dict = {c: (lambda v: ru_rub(v, 0)) for c in fallback.columns if c not in skip_rub}
        fmt_dict["Комиссия, %"] = ru_pct
        fmt_dict["ROI операц., %"] = ru_pct
        fmt_dict["ROI чистая, %"] = ru_pct
        st.dataframe(fallback.style.format(fmt_dict, na_rep="—"), use_container_width=True, hide_index=True)

    st.divider()
    st.caption(
        "**Себестоимость** — нетто по возвратам. **Маркетинг** = реклама по SKU (Seller API) + "
        "Performance API CPC/CPO (если подключён и включён) + магазинная реклама, не заменённая "
        "Performance API — без задвоения. **ROI** считается от себестоимости (юнит-экономика), "
        "а не от выручки — если нужна другая формула, легко поменять. История периодов пока не "
        "сохраняется между загрузками — это отдельная задача (Google Sheets)."
    )

# ── Навигация (левое меню, как в стороннем сервисе) ──────────────────────────
# Диаграммы и Калькулятор убраны по просьбе пользователя (02.08.2026) — раньше были
# отдельными вкладками st.tabs(), больше нигде не используются.
pg = st.navigation([
    st.Page(page_articles, title="По артикулам", icon="📋", default=True),
    st.Page(page_pnl,      title="P&L",          icon="📈"),
    st.Page(page_stocks,   title="Остатки",      icon="📦"),
    st.Page(page_details,  title="Детализация",  icon="🔍"),
])
pg.run()